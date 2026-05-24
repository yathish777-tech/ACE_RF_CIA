import io
from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user
from models import db, RetestApplication, SeatingAllotment, ExamAttendance
from datetime import datetime
from functools import wraps

tutor_bp = Blueprint('tutor', __name__)

def tutor_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not (current_user.role == 'tutor' or current_user.secondary_role == 'tutor'):
            flash('Access denied.', 'danger')
            return redirect(url_for('main.index'))
        return f(*args, **kwargs)
    return decorated


@tutor_bp.route('/dashboard')
@login_required
@tutor_required
def dashboard():
    year_filter    = request.args.get('year', type=int)
    section_filter = request.args.get('section', '')

    pending_q = RetestApplication.query.filter_by(
        tutor_id=current_user.id, staff_status='approved', tutor_status='pending')
    reviewed_q = RetestApplication.query.filter(
        RetestApplication.tutor_id == current_user.id,
        RetestApplication.tutor_status.in_(['approved', 'rejected']))

    if year_filter:
        pending_q  = pending_q.filter_by(student_year=year_filter)
        reviewed_q = reviewed_q.filter_by(student_year=year_filter)
    if section_filter:
        pending_q  = pending_q.filter_by(student_section=section_filter)
        reviewed_q = reviewed_q.filter_by(student_section=section_filter)

    pending  = pending_q.order_by(RetestApplication.submitted_at.desc()).all()
    reviewed = reviewed_q.order_by(RetestApplication.tutor_action_time.desc()).all()

    stats = {
        'pending' : len(pending),
        'approved': sum(1 for r in reviewed if r.tutor_status == 'approved'),
        'rejected': sum(1 for r in reviewed if r.tutor_status == 'rejected'),
        'total'   : len(pending) + len(reviewed),
    }
    return render_template('tutor/dashboard.html',
                           pending=pending, reviewed=reviewed, stats=stats,
                           year_filter=year_filter, section_filter=section_filter,
                           sections=['A', 'B', 'C'])


@tutor_bp.route('/action/<int:app_id>', methods=['POST'])
@login_required
@tutor_required
def action(app_id):
    application = RetestApplication.query.filter_by(
        id=app_id, tutor_id=current_user.id).first_or_404()
    act    = request.form.get('action')
    remark = request.form.get('remark', '').strip()
    application.tutor_status = 'approved' if act == 'approve' else 'rejected'
    application.tutor_remark = remark
    application.tutor_action_time = datetime.utcnow()
    if act == 'reject':
        application.final_status = 'rejected'
        db.session.commit()
        try:
            from utils.email_utils import notify_student_final
            notify_student_final(application)
        except: pass
        flash('Application rejected.', 'warning')
    else:
        db.session.commit()
        try:
            if application.submission_type == 'pre':
                from utils.email_utils import notify_hod_after_tutor_pre
                notify_hod_after_tutor_pre(application)
            else:
                from utils.email_utils import notify_coordinator_after_tutor_late
                notify_coordinator_after_tutor_late(application)
        except: pass
        flash('Application approved and forwarded.', 'success')
    return redirect(url_for('tutor.dashboard'))


# ─── DOWNLOAD ─────────────────────────────────────────────────────────────────
@tutor_bp.route('/applications/download/<fmt>')
@login_required
@tutor_required
def download_applications(fmt):
    year_filter    = request.args.get('year', type=int)
    section_filter = request.args.get('section', '')

    query = RetestApplication.query.filter_by(tutor_id=current_user.id)
    if year_filter:    query = query.filter_by(student_year=year_filter)
    if section_filter: query = query.filter_by(student_section=section_filter)
    apps = query.order_by(RetestApplication.submitted_at.desc()).all()

    rows = [{'ID': a.id, 'Student': a.student_name, 'Reg No': a.register_number,
             'Year': a.student_year or '', 'Section': a.student_section or '',
             'Subject': a.subject.subject_name, 'Semester': a.semester,
             'CIA No': a.cia_number, 'Type': a.submission_type.upper(),
             'Staff': a.staff_status, 'Tutor': a.tutor_status,
             'HOD': a.hod_status, 'Final': a.final_status,
             'Submitted': a.submitted_at.strftime('%d %b %Y %H:%M')} for a in apps]

    label = ''
    if year_filter:    label += f'_Year{year_filter}'
    if section_filter: label += f'_Sec{section_filter}'

    if fmt == 'excel':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'My Applications'
        if rows:
            hdrs = list(rows[0].keys())
            for ci, h in enumerate(hdrs, 1):
                c = ws.cell(1, ci, h)
                c.font = Font(bold=True, color='FFFFFF', name='Arial')
                c.fill = PatternFill('solid', fgColor='1A237E')
                c.alignment = Alignment(horizontal='center')
            for ri, row in enumerate(rows, 2):
                for ci, v in enumerate(row.values(), 1): ws.cell(ri, ci, v)
                bg = ('E8F5E9' if row['Final']=='approved' else
                      'FFEBEE' if row['Final']=='rejected' else 'FFFFFF')
                for ci in range(1, len(hdrs)+1):
                    ws.cell(ri, ci).fill = PatternFill('solid', fgColor=bg)
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 15
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf, as_attachment=True,
            download_name=f'tutor_applications{label}_{datetime.now().strftime("%Y%m%d")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=20, rightMargin=20, topMargin=30, bottomMargin=20)
        styles = getSampleStyleSheet()
        title_str = f'Tutor Applications — {current_user.name}'
        if year_filter:    title_str += f' | Year {year_filter}'
        if section_filter: title_str += f' | Section {section_filter}'
        els = [Paragraph(title_str, styles['Title']),
               Paragraph(f'Generated: {datetime.now().strftime("%d %b %Y %H:%M")}', styles['Normal']),
               Spacer(1, 12)]
        if rows:
            hdrs = ['#','Student','Reg No','Yr','Sec','Subject','Sem','CIA','Type','Tutor','HOD','Final']
            td = [hdrs]+[[str(r['ID']),r['Student'][:14],r['Reg No'],str(r['Year']),
                           str(r['Section']),r['Subject'][:14],str(r['Semester']),
                           str(r['CIA No']),r['Type'],r['Tutor'].upper(),
                           r['HOD'].upper(),r['Final'].upper()] for r in rows]
            t = Table(td, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1A237E')),
                ('TEXTCOLOR',(0,0),(-1,0),colors.white),
                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
                ('FONTSIZE',(0,0),(-1,-1),8),
                ('GRID',(0,0),(-1,-1),0.5,colors.grey),
                ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F5F7FF')]),
                ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ]))
            els.append(t)
        doc.build(els); buf.seek(0)
        return send_file(buf, as_attachment=True,
            download_name=f'tutor_applications{label}_{datetime.now().strftime("%Y%m%d")}.pdf',
            mimetype='application/pdf')


# ─── RETRANSMIT (Tutor) ──────────────────────────────────────────────────────
@tutor_bp.route('/applications/<int:app_id>/retransmit', methods=['POST'])
@login_required
@tutor_required
def retransmit(app_id):
    """Tutor can retransmit a mistakenly-rejected application from the rejected stage."""
    app = RetestApplication.query.filter_by(
        id=app_id, tutor_id=current_user.id).first_or_404()

    if app.final_status != 'rejected':
        flash('Only rejected applications can be retransmitted.', 'warning')
        return redirect(url_for('tutor.dashboard'))

    # Reset only the stage that was rejected
    if app.hod_status == 'rejected':
        app.hod_status = 'pending'; app.hod_remark = None; app.hod_action_time = None
    elif app.coordinator_status == 'rejected':
        app.coordinator_status = 'pending'; app.coordinator_remark = None; app.coordinator_action_time = None
    elif app.tutor_status == 'rejected':
        app.tutor_status = 'pending'; app.tutor_remark = None; app.tutor_action_time = None
    elif app.staff_status == 'rejected':
        app.staff_status = 'pending'; app.staff_remark = None; app.staff_action_time = None

    app.final_status     = 'pending'
    app.retransmit_count = (app.retransmit_count or 0) + 1
    app.retransmit_by    = current_user.id
    app.retransmit_at    = datetime.utcnow()
    db.session.commit()
    flash(f'Application #{app_id} retransmitted successfully.', 'success')
    return redirect(url_for('tutor.dashboard'))


# ─── MY HALLS (Tutor as invigilator) ────────────────────────────────────────
@tutor_bp.route('/my-halls')
@login_required
@tutor_required
def my_halls():
    halls = SeatingAllotment.query.filter(
        (SeatingAllotment.invigilator_id == None) | (SeatingAllotment.invigilator_id == current_user.id)
    ).order_by(SeatingAllotment.hall_number).all()
    hall_numbers = sorted({h.hall_number for h in halls})
    hall_status = {}
    for hn in hall_numbers:
        rows   = SeatingAllotment.query.filter(
            SeatingAllotment.hall_number == hn,
            (SeatingAllotment.invigilator_id == None) | (SeatingAllotment.invigilator_id == current_user.id)
        ).all()
        total  = sum(len(r.get_register_numbers()) for r in rows)
        marked = ExamAttendance.query.filter_by(hall_number=hn, marked_by=current_user.id).count()
        hall_status[hn] = {'total': total, 'marked': marked}
    return render_template('admin/my_halls.html',
                           halls=halls, hall_numbers=hall_numbers, hall_status=hall_status)
