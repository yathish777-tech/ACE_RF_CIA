import os, io, uuid, re
from flask import (Blueprint, render_template, redirect, url_for,
                   flash, request, jsonify, current_app, send_file)
from flask_login import login_required, current_user
from models import db, User, Subject, CIADate, RetestApplication, AbsenceRecord, SubjectStaffSection, SeatingAllotment, ExamAttendance, Hall, SeatingAllocation, HallAttendance
from datetime import datetime, date, timedelta
from functools import wraps

admin_bp = Blueprint('admin', __name__)

SEMESTER_TO_YEAR = {1: 1, 2: 1, 3: 2, 4: 2, 5: 3, 6: 3, 7: 4, 8: 4}
SECTIONS = ['A', 'B', 'C']
STUDENT_DEFAULT_PASSWORD = 'student123'
STUDENT_EMAIL_DOMAIN = 'student.local'

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if current_user.role != 'admin':
            flash('Access denied.', 'danger')
            return redirect(url_for('main.index'))
        return f(*args, **kwargs)
    return decorated


def admin_or_hod_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not (current_user.role in ('admin', 'hod') or current_user.secondary_role == 'hod'):
            flash('Access denied.', 'danger')
            return redirect(url_for('main.index'))
        return f(*args, **kwargs)
    return decorated


# ─── DASHBOARD ──────────────────────────────────────────────────────────────
@admin_bp.route('/dashboard')
@login_required
@admin_required
def dashboard():
    total_apps     = RetestApplication.query.count()
    approved       = RetestApplication.query.filter_by(final_status='approved').count()
    rejected       = RetestApplication.query.filter_by(final_status='rejected').count()
    pending        = RetestApplication.query.filter_by(final_status='pending').count()
    pre_count      = RetestApplication.query.filter_by(submission_type='pre').count()
    late_count     = RetestApplication.query.filter_by(submission_type='late').count()
    total_students = User.query.filter_by(role='student').count()
    total_staff    = User.query.filter(
        User.role.in_(['subject_staff','tutor','hod','coordinator'])).count()

    subjects = Subject.query.filter_by(is_active=True).all()
    subject_stats = sorted([
        {'name': s.subject_name, 'code': s.subject_code,
         'count': RetestApplication.query.filter_by(subject_id=s.id).count()}
        for s in subjects], key=lambda x: x['count'], reverse=True)

    recent_apps = RetestApplication.query\
        .order_by(RetestApplication.submitted_at.desc()).limit(10).all()

    # Year-wise section-wise stats
    year_section_stats = _get_year_section_stats()

    stats = {'total': total_apps, 'approved': approved, 'rejected': rejected,
             'pending': pending, 'pre': pre_count, 'late': late_count,
             'students': total_students, 'staff': total_staff}
    return render_template('admin/dashboard.html', stats=stats,
                           subject_stats=subject_stats, recent_apps=recent_apps,
                           year_section_stats=year_section_stats, sections=SECTIONS)


def _get_year_section_stats():
    """Returns year->section->list of applications mapping."""
    apps = RetestApplication.query.all()
    stats = {}
    for year in range(1, 5):
        stats[year] = {}
        for sec in SECTIONS:
            year_apps = [a for a in apps
                         if (a.student_year == year and a.student_section == sec)]
            stats[year][sec] = {
                'total': len(year_apps),
                'approved': sum(1 for a in year_apps if a.final_status == 'approved'),
                'rejected': sum(1 for a in year_apps if a.final_status == 'rejected'),
                'pending': sum(1 for a in year_apps if a.final_status == 'pending'),
                'apps': year_apps
            }
    return stats


def _clean_cell(value) -> str:
    if value is None:
        return ''
    text = str(value).strip()
    if text.lower() in ('nan', 'none', 'na', ''):
        return ''
    if re.fullmatch(r'\d+\.0', text):
        text = text[:-2]
    return text.strip()


def _normalize_column_name(col) -> str:
    return re.sub(r'_+', '_', re.sub(r'[^a-z0-9]+', '_', str(col).strip().lower())).strip('_')


def _normalize_register_number(value) -> str:
    return re.sub(r'\s+', '', _clean_cell(value)).upper()


def _parse_year(value):
    text = _clean_cell(value)
    if not text:
        return None
    try:
        year = int(float(text))
    except ValueError:
        return None
    return year if year in (1, 2, 3, 4) else None


def _parse_section(value):
    section = _clean_cell(value).upper()
    return section if section in SECTIONS else None


def _parse_date(value):
    text = _clean_cell(value)
    if not text:
        return None
    try:
        import pandas as pd
        return pd.to_datetime(text, dayfirst=True).date()
    except Exception:
        try:
            return datetime.strptime(text, '%Y-%m-%d').date()
        except ValueError:
            return None


def _row_value(row, *keys):
    for key in keys:
        value = row.get(key, '')
        value = _clean_cell(value)
        if value:
            return value
    return ''


def _find_student_by_register(register_number):
    reg = _normalize_register_number(register_number)
    if not reg:
        return None
    return User.query.filter(
        User.role == 'student',
        db.func.upper(User.register_number) == reg
    ).first()


def _generated_student_email(register_number):
    safe = re.sub(r'[^a-z0-9._-]+', '_', register_number.lower()).strip('._-') or uuid.uuid4().hex
    candidate = f'{safe}@{STUDENT_EMAIL_DOMAIN}'
    suffix = 2
    while User.query.filter(db.func.lower(User.email) == candidate.lower()).first():
        candidate = f'{safe}{suffix}@{STUDENT_EMAIL_DOMAIN}'
        suffix += 1
    return candidate


def _validate_student_payload(register_number, name, year, section):
    errors = []
    if not register_number:
        errors.append('Register number is required.')
    if not name:
        errors.append('Student name is required.')
    if year not in (1, 2, 3, 4):
        errors.append('Year must be 1, 2, 3, or 4.')
    if section not in SECTIONS:
        errors.append(f'Section must be one of {", ".join(SECTIONS)}.')
    return errors


def _upsert_student_record(register_number, name, year, section, email='', phone=''):
    register_number = _normalize_register_number(register_number)
    name = _clean_cell(name)
    email = _clean_cell(email).lower()
    phone = _clean_cell(phone)

    print("=" * 80)
    print("Processing Student")
    print(f"Register Number : {register_number}")
    print(f"Name            : {name}")
    print(f"Year            : {year}")
    print(f"Section         : {section}")
    print(f"Email           : {email}")
    print(f"Phone           : {phone}")

    errors = _validate_student_payload(register_number, name, year, section)
    if errors:
        print("VALIDATION FAILED:", errors)
        return 'skipped', '; '.join(errors)

    existing = _find_student_by_register(register_number)
    email_user = None

    if email:
        email_user = User.query.filter(
            db.func.lower(User.email) == email
        ).first()

        if email_user and email_user.role != 'student':
            print("SKIPPED: Email belongs to staff/admin")
            return 'skipped', f'Email {email} belongs to a staff/admin account.'

        if existing and email_user and email_user.id != existing.id:
            print("SKIPPED: Email belongs to another student")
            return 'skipped', f'Email {email} belongs to another student.'

        if not existing:
            existing = email_user

    if existing:
        print("Updating existing student...")

        existing.name = name
        existing.register_number = register_number
        existing.year = year
        existing.section = section

        if email:
            existing.email = email

        if phone:
            existing.phone = phone

        existing.is_active = True

        print("UPDATED SUCCESSFULLY")
        return 'updated', ''

    print("Creating new student...")

    student = User(
        name=name,
        email=email or _generated_student_email(register_number),
        phone=phone,
        role='student',
        register_number=register_number,
        year=year,
        section=section,
        is_active=True
    )

    student.set_password(STUDENT_DEFAULT_PASSWORD)

    db.session.add(student)

    print("ADDED TO SESSION")
    print("=" * 80)

    return 'added', ''


def _allotment_student_count(allotment):
    regs = allotment.get_register_numbers()
    return len(regs) if regs else (allotment.total_students or allotment.num_students or 0)


def _absence_student_year(record, student=None):
    if student:
        value = student.get('year') or student.get('yr')
        try:
            if value:
                year = int(float(value))
                if year in (1, 2, 3, 4):
                    return year
        except (TypeError, ValueError):
            pass
    if record.semester:
        return SEMESTER_TO_YEAR.get(record.semester)
    if record.uploader and record.uploader.handling_year:
        return record.uploader.handling_year
    return None


def _absence_records_for_cia(cia_number, year_filter=None):
    records = AbsenceRecord.query.filter_by(cia_number=cia_number)\
        .order_by(AbsenceRecord.uploaded_at.desc()).all()
    if not year_filter:
        return records
    filtered = []
    for record in records:
        students = record.get_students()
        if students:
            if any(_absence_student_year(record, student) == year_filter for student in students):
                filtered.append(record)
        elif _absence_student_year(record) == year_filter:
            filtered.append(record)
    return filtered


# STUDENT REGISTRY
@admin_bp.route('/students')
@login_required
@admin_required
def manage_students():
    year_filter = request.args.get('year', type=int)
    section_filter = (request.args.get('section', '') or '').upper()
    search = (request.args.get('q', '') or '').strip()

    query = User.query.filter_by(role='student')
    if year_filter:
        query = query.filter_by(year=year_filter)
    if section_filter:
        query = query.filter_by(section=section_filter)
    if search:
        like = f'%{search}%'
        query = query.filter(
            User.name.ilike(like) |
            User.register_number.ilike(like) |
            User.email.ilike(like)
        )

    students = query.order_by(User.year, User.section, User.register_number, User.name).all()
    total_students = User.query.filter_by(role='student').count()
    active_students = User.query.filter_by(role='student', is_active=True).count()
    return render_template('admin/manage_students.html',
                           students=students, sections=SECTIONS,
                           year_filter=year_filter,
                           section_filter=section_filter,
                           search=search,
                           total_students=total_students,
                           active_students=active_students,
                           default_password=STUDENT_DEFAULT_PASSWORD)


@admin_bp.route('/students/upload', methods=['POST'])
@login_required
@admin_required
def upload_students():
    f = request.files.get('student_file')

    if not f or not f.filename:
        flash('Please select a Student Details file.', 'danger')
        return redirect(url_for('admin.manage_students'))

    ext = f.filename.rsplit('.', 1)[-1].lower()

    if ext not in ('xlsx', 'xls', 'csv'):
        flash('Only Excel (.xlsx/.xls) or CSV files accepted.', 'danger')
        return redirect(url_for('admin.manage_students'))

    added = 0
    updated = 0
    skipped = 0

    try:
        import pandas as pd

        # Read file
        if ext == 'csv':
            df = pd.read_csv(f, dtype=str)
        else:
            df = pd.read_excel(f, dtype=str)

        # Normalize column names
        df.columns = [_normalize_column_name(c) for c in df.columns]

        print("\n================ STUDENT UPLOAD STARTED ================\n")
        print("Columns Found:", df.columns.tolist())
        print("\nFirst 5 Rows:")
        print(df.head())
        print("\n========================================================\n")

        for index, row in df.iterrows():

            register_number = _normalize_register_number(
                _row_value(
                    row,
                    'register_number',
                    'registernumber',
                    'register_no',
                    'reg_no',
                    'reg',
                    'roll_no'
                )
            )

            name = _row_value(
                row,
                'name',
                'student_name',
                'student'
            )

            year = _parse_year(
                _row_value(
                    row,
                    'year',
                    'yr'
                )
            )

            section = _parse_section(
                _row_value(
                    row,
                    'section',
                    'sec'
                )
            )

            email = _row_value(
                row,
                'email',
                'email_id',
                'student_email',
                'student_email_id'
            )

            phone = _row_value(
                row,
                'phone',
                'phone_number',
                'mobile'
            )

            print("\n--------------------------------------------------")
            print(f"Row             : {index + 1}")
            print(f"Register Number : {register_number}")
            print(f"Name            : {name}")
            print(f"Year            : {year}")
            print(f"Section         : {section}")
            print(f"Email           : {email}")
            print(f"Phone           : {phone}")

            status, message = _upsert_student_record(
                register_number,
                name,
                year,
                section,
                email,
                phone
            )

            print(f"Status          : {status}")
            print(f"Message         : {message}")
            print("--------------------------------------------------")

            if status == 'added':
                added += 1

            elif status == 'updated':
                updated += 1

            else:
                skipped += 1

        print("\n================ COMMITTING DATABASE ================\n")
        db.session.commit()
        print("DATABASE COMMIT SUCCESSFUL")
        print(f"Added   : {added}")
        print(f"Updated : {updated}")
        print(f"Skipped : {skipped}")
        print("\n=====================================================\n")

        flash(
            f'Student upload complete: '
            f'{added} added, '
            f'{updated} updated, '
            f'{skipped} skipped. '
            f'Default password for new accounts: {STUDENT_DEFAULT_PASSWORD}',
            'success'
        )

    except Exception as e:
        db.session.rollback()
        print("\n================ ERROR =================")
        import traceback
        traceback.print_exc()
        print("=========================================\n")
        flash(f'Student upload error: {e}', 'danger')

    return redirect(url_for('admin.manage_students'))

@admin_bp.route('/students/save', methods=['POST'])
@login_required
@admin_required
def save_student():
    student_id = request.form.get('student_id', '').strip()
    register_number = _normalize_register_number(request.form.get('register_number', ''))
    name = _clean_cell(request.form.get('name', ''))
    year = _parse_year(request.form.get('year', ''))
    section = _parse_section(request.form.get('section', ''))
    email = _clean_cell(request.form.get('email', '')).lower()
    phone = _clean_cell(request.form.get('phone', ''))
    password = request.form.get('password', '').strip()

    errors = _validate_student_payload(register_number, name, year, section)
    if errors:
        flash(' '.join(errors), 'danger')
        return redirect(url_for('admin.manage_students'))

    try:
        if student_id:
            student = User.query.filter_by(id=int(student_id), role='student').first_or_404()
            duplicate = _find_student_by_register(register_number)
            if duplicate and duplicate.id != student.id:
                flash('Another student already uses this register number.', 'danger')
                return redirect(url_for('admin.manage_students'))
            if email:
                email_user = User.query.filter(db.func.lower(User.email) == email).first()
                if email_user and email_user.id != student.id:
                    flash('Another account already uses this email.', 'danger')
                    return redirect(url_for('admin.manage_students'))
                student.email = email
            student.name = name
            student.register_number = register_number
            student.year = year
            student.section = section
            student.phone = phone
            student.is_active = True
            if password:
                student.set_password(password)
            db.session.commit()
            flash('Student updated.', 'success')
        else:
            status, message = _upsert_student_record(register_number, name, year, section, email, phone)
            if status == 'skipped':
                flash(message, 'danger')
                return redirect(url_for('admin.manage_students'))
            if password:
                student = _find_student_by_register(register_number)
                if student:
                    student.set_password(password)
            db.session.commit()
            flash(f'Student {status}. Default password: {password or STUDENT_DEFAULT_PASSWORD}', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Student save error: {e}', 'danger')
    return redirect(url_for('admin.manage_students'))


@admin_bp.route('/students/<int:uid>/toggle', methods=['POST'])
@login_required
@admin_required
def toggle_student(uid):
    student = User.query.filter_by(id=uid, role='student').first_or_404()
    student.is_active = not student.is_active
    db.session.commit()
    flash(f'Student {"activated" if student.is_active else "deactivated"}.', 'success')
    return redirect(url_for('admin.manage_students'))


# ─── BULK UPLOAD PAGE (GET) ──────────────────────────────────────────────────
@admin_bp.route('/bulk-upload', methods=['GET'])
@login_required
@admin_required
def bulk_upload():
    subjects  = Subject.query.order_by(Subject.subject_name).all()
    cia_dates = CIADate.query.order_by(CIADate.exam_date.desc()).all()
    staff_list = User.query.filter(
        User.role.in_(['subject_staff','tutor','hod','coordinator'])
    ).order_by(User.name).all()
    sss_list = SubjectStaffSection.query.order_by(
        SubjectStaffSection.semester, SubjectStaffSection.section).all()
    return render_template('admin/bulk_upload.html',
                           subjects=subjects, cia_dates=cia_dates,
                           staff_list=staff_list, sss_list=sss_list)


# ─── UPLOAD 1: STAFF DETAILS FILE ───────────────────────────────────────────
@admin_bp.route('/bulk-upload/staff', methods=['POST'])
@login_required
@admin_required
def bulk_upload_staff():
    """
    Upload Staff Details Excel/CSV.
    Required columns: staff_name, staff_email, role, phone, handling_year, handling_section
    """
    f = request.files.get('staff_file')
    if not f or not f.filename:
        flash('Please select a Staff Details file.', 'danger')
        return redirect(url_for('admin.bulk_upload'))
    ext = f.filename.rsplit('.', 1)[-1].lower()
    if ext not in ('xlsx','xls','csv'):
        flash('Only Excel (.xlsx/.xls) or CSV files accepted.', 'danger')
        return redirect(url_for('admin.bulk_upload'))
    try:
        import pandas as pd
        df = pd.read_csv(f) if ext == 'csv' else pd.read_excel(f, dtype=str)
        df.columns = [c.strip().lower().replace(' ', '_') for c in df.columns]

        added = updated = 0
        VALID_ROLES = ('subject_staff', 'tutor', 'hod', 'coordinator')

        for _, row in df.iterrows():
            row = {k: (str(v).strip() if str(v).strip().lower() not in ('nan','none','') else '')
                   for k, v in row.items()}
            name  = row.get('staff_name') or row.get('name', '')
            email = (row.get('staff_email') or row.get('email', '')).lower()
            role  = (row.get('role') or row.get('staff_role', 'subject_staff'))\
                    .strip().lower().replace(' ', '_')
            dept  = row.get('department', '')
            phone = row.get('phone', '') or row.get('phone_number', '')
            h_year = row.get('handling_year', '')
            h_sec  = (row.get('handling_section', '') or '').upper().strip()

            if not email or not name:
                continue
            if role not in VALID_ROLES:
                role = 'subject_staff'

            h_year_int = int(float(h_year)) if h_year else None
            h_sec_val = h_sec if h_sec in ('A','B','C') else None

            existing = User.query.filter(db.func.lower(User.email) == email).first()
            if existing:
                if existing.role != role and not existing.secondary_role:
                    existing.secondary_role = role
                if dept:  existing.department = dept
                if phone: existing.phone = phone
                if h_year_int: existing.handling_year = h_year_int
                if h_sec_val:  existing.handling_section = h_sec_val
                updated += 1
            else:
                u = User(name=name, email=email, role=role,
                         department=dept, phone=phone,
                         handling_year=h_year_int, handling_section=h_sec_val)
                u.set_password('staff123')
                db.session.add(u)
                added += 1

        db.session.commit()
        flash(f'Staff upload complete: {added} added, {updated} updated. '
              f'Default password for new accounts: staff123', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Staff upload error: {e}', 'danger')
    return redirect(url_for('admin.bulk_upload'))


# ─── UPLOAD 2: SUBJECT FILE ──────────────────────────────────────────────────
@admin_bp.route('/bulk-upload/subjects', methods=['POST'])
@login_required
@admin_required
def bulk_upload_subjects():
    """
    Upload Subject Details Excel/CSV.
    Columns: subject_name, subject_code, semester, section, shsn (staff email or name),
             department
    Each row = one subject+section+staff combo.
    """
    f = request.files.get('subject_file')
    if not f or not f.filename:
        flash('Please select a Subject file.', 'danger')
        return redirect(url_for('admin.bulk_upload'))
    ext = f.filename.rsplit('.', 1)[-1].lower()
    if ext not in ('xlsx','xls','csv'):
        flash('Only Excel (.xlsx/.xls) or CSV files accepted.', 'danger')
        return redirect(url_for('admin.bulk_upload'))
    try:
        import pandas as pd
        df = pd.read_csv(f) if ext == 'csv' else pd.read_excel(f, dtype=str)
        df.columns = [c.strip().lower().replace(' ', '_') for c in df.columns]

        added_sub = added_map = updated_map = 0

        for _, row in df.iterrows():
            row = {k: (str(v).strip() if str(v).strip().lower() not in ('nan','none','') else '')
                   for k, v in row.items()}
            sub_name  = row.get('subject_name') or row.get('subject', '')
            sub_code  = row.get('subject_code') or row.get('code', '')
            semester  = row.get('semester', '')
            section   = (row.get('section', '') or '').upper().strip()
            shsn      = row.get('shsn') or row.get('staff_email') or row.get('staff_name', '')
            dept      = row.get('department', '')

            if not (sub_name and sub_code and semester):
                continue

            sem_int = int(float(semester))
            year    = SEMESTER_TO_YEAR.get(sem_int, 1)
            sec_val = section if section in ('A','B','C') else None

            # Find/create subject
            subject = Subject.query.filter_by(
                subject_code=sub_code, semester=sem_int).first()
            if not subject:
                subject = Subject(subject_name=sub_name, subject_code=sub_code,
                                  semester=sem_int, department=dept, year=year)
                db.session.add(subject)
                db.session.flush()
                added_sub += 1
            else:
                if dept: subject.department = dept

            # Resolve staff by email or name
            staff_obj = None
            if shsn:
                if '@' in shsn:
                    staff_obj = User.query.filter(
                        db.func.lower(User.email) == shsn.lower()).first()
                else:
                    staff_obj = User.query.filter(
                        db.func.lower(User.name) == shsn.lower()).first()

            # Set default staff on subject
            if staff_obj and not subject.staff_id:
                subject.staff_id = staff_obj.id

            # Create section mapping if section specified
            if sec_val and staff_obj:
                existing_sss = SubjectStaffSection.query.filter_by(
                    subject_id=subject.id, section=sec_val, semester=sem_int).first()
                if existing_sss:
                    existing_sss.staff_id = staff_obj.id
                    updated_map += 1
                else:
                    db.session.add(SubjectStaffSection(
                        subject_id=subject.id, staff_id=staff_obj.id,
                        semester=sem_int, section=sec_val))
                    added_map += 1

        db.session.commit()
        flash(f'Subject upload complete: {added_sub} subjects created, '
              f'{added_map} section mappings added, {updated_map} updated.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Subject upload error: {e}', 'danger')
    return redirect(url_for('admin.bulk_upload'))


# ─── UPLOAD 3: CIA SCHEDULE FILE ─────────────────────────────────────────────
@admin_bp.route('/bulk-upload/cia', methods=['POST'])
@login_required
@admin_required
def bulk_upload_cia():
    """
    Upload CIA Schedule Excel/CSV.
    Required: subject_code, semester, cia_number, cia_exam_date
    Optional: cia_deadline_date, cia_retest_date, academic_year, section
    """
    f = request.files.get('cia_file')
    if not f or not f.filename:
        flash('Please select a CIA Schedule file.', 'danger')
        return redirect(url_for('admin.bulk_upload'))
    ext = f.filename.rsplit('.', 1)[-1].lower()
    if ext not in ('xlsx','xls','csv'):
        flash('Only Excel (.xlsx/.xls) or CSV files accepted.', 'danger')
        return redirect(url_for('admin.bulk_upload'))
    try:
        import pandas as pd
        df = pd.read_csv(f) if ext == 'csv' else pd.read_excel(f, dtype=str)
        df.columns = [c.strip().lower().replace(' ', '_') for c in df.columns]

        added_cia = updated_cia = 0

        for _, row in df.iterrows():
            row = {k: (str(v).strip() if str(v).strip().lower() not in ('nan','none','') else '')
                   for k, v in row.items()}

            sub_code   = row.get('subject_code') or row.get('code', '')
            sub_name   = row.get('subject_name') or row.get('subject', '')
            semester   = row.get('semester', '')
            cia_num    = row.get('cia_number') or row.get('cia_no', '')
            exam_str   = row.get('cia_exam_date') or row.get('exam_date', '')
            retest_str = row.get('cia_retest_date') or row.get('retest_date', '')
            dead_str   = row.get('cia_deadline_date') or row.get('deadline', '') or row.get('cia_deadline', '')
            acad_year  = row.get('academic_year', '')

            if not (semester and cia_num and exam_str):
                continue

            sem_int = int(float(semester))
            cia_int = int(float(cia_num))

            # Find subject
            subject = None
            if sub_code:
                subject = Subject.query.filter_by(
                    subject_code=sub_code, semester=sem_int).first()
            if not subject and sub_name:
                subject = Subject.query.filter_by(
                    subject_name=sub_name, semester=sem_int).first()
            if not subject:
                continue

            try:
                ed = pd.to_datetime(exam_str).date()
                rd = pd.to_datetime(retest_str).date() if retest_str else None
                nd = pd.to_datetime(dead_str).date()   if dead_str   else None
            except Exception:
                continue

            existing = CIADate.query.filter_by(
                subject_id=subject.id, cia_number=cia_int).first()
            if existing:
                existing.exam_date = ed
                if rd: existing.retest_date = rd
                if nd: existing.application_end_date = nd
                if acad_year: existing.academic_year = acad_year
                updated_cia += 1
            else:
                db.session.add(CIADate(
                    subject_id=subject.id, cia_number=cia_int,
                    exam_date=ed, retest_date=rd,
                    application_end_date=nd, semester=sem_int,
                    academic_year=acad_year, created_by=current_user.id))
                added_cia += 1

        db.session.commit()
        flash(f'CIA Schedule upload complete: {added_cia} added, {updated_cia} updated.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'CIA upload error: {e}', 'danger')
    return redirect(url_for('admin.bulk_upload'))


# ─── PORTAL WINDOW CONTROLS ──────────────────────────────────────────────────
@admin_bp.route('/bulk-upload/cia-dates/<int:cid>/toggle-window', methods=['POST'])
@login_required
def toggle_retest_window_legacy(cid):
    allowed = (current_user.role == 'admin' or
               current_user.role == 'hod' or
               current_user.secondary_role == 'hod')
    if not allowed:
        flash('Access denied.', 'danger')
        return redirect(url_for('main.index'))
    cia = CIADate.query.get_or_404(cid)
    today = date.today()
    if cia.is_application_open():
        cia.application_end_date = today - timedelta(days=1)
    else:
        if not cia.exam_date or cia.exam_date >= today:
            cia.exam_date = today - timedelta(days=1)
        cia.application_end_date = today + timedelta(days=7)
    db.session.commit()
    state = 'OPENED' if cia.is_application_open() else 'CLOSED'
    flash(f'Portal window {state} for {cia.subject.subject_name} — CIA {cia.cia_number}.', 'success')
    return redirect(request.referrer or url_for('admin.manage_cia_dates'))


@admin_bp.route('/cia-dates/<int:cid>/set-window', methods=['POST'])
@login_required
def set_retest_window(cid):
    allowed = (current_user.role == 'admin' or
               current_user.role == 'hod' or
               current_user.secondary_role == 'hod')
    if not allowed:
        flash('Access denied.', 'danger')
        return redirect(url_for('main.index'))
    cia = CIADate.query.get_or_404(cid)
    open_date = request.form.get('open_until_date', '').strip()
    try:
        new_end = datetime.strptime(open_date, '%Y-%m-%d').date()
        cia.application_end_date    = new_end
        db.session.commit()
        flash(f'Portal reopened for {cia.subject.subject_name} CIA {cia.cia_number}. '
              f'New deadline: {new_end.strftime("%d %b %Y")}.', 'success')
    except ValueError:
        flash('Invalid date format.', 'danger')
    return redirect(request.referrer or url_for('admin.manage_cia_dates'))


# ─── DOWNLOAD APPLICATIONS (year/section filter) ─────────────────────────────
@admin_bp.route('/applications/download/<fmt>')
@login_required
@admin_required
def download_applications(fmt):
    year_filter    = request.args.get('year', type=int)
    section_filter = request.args.get('section', '')

    query = RetestApplication.query
    if year_filter:
        query = query.filter_by(student_year=year_filter)
    if section_filter:
        query = query.filter_by(student_section=section_filter)
    apps = query.order_by(RetestApplication.submitted_at.desc()).all()

    rows = [{'ID': a.id, 'Student': a.student_name, 'Reg No': a.register_number,
             'Email': a.student_email,
             'Year': a.student_year or '', 'Section': a.student_section or '',
             'Subject': a.subject.subject_name, 'Code': a.subject.subject_code,
             'Semester': a.semester, 'CIA No': a.cia_number, 'CIA Date': str(a.cia_date),
             'Reason': a.reason_type.replace('_', ' ').title(),
             'Type': a.submission_type.upper(),
             'Staff': a.staff_status, 'Tutor': a.tutor_status,
             'Coordinator': a.coordinator_status, 'HOD': a.hod_status,
             'Final': a.final_status,
             'Submitted': a.submitted_at.strftime('%d %b %Y %H:%M')} for a in apps]

    label = ''
    if year_filter: label += f'_Year{year_filter}'
    if section_filter: label += f'_Sec{section_filter}'

    if fmt == 'excel':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Applications'
        if rows:
            hdrs = list(rows[0].keys())
            for ci, h in enumerate(hdrs, 1):
                c = ws.cell(1, ci, h)
                c.font = Font(bold=True, color='FFFFFF', name='Arial')
                c.fill = PatternFill('solid', fgColor='1A237E')
                c.alignment = Alignment(horizontal='center')
            for ri, row in enumerate(rows, 2):
                for ci, v in enumerate(row.values(), 1):
                    ws.cell(ri, ci, v)
                bg = ('E8F5E9' if row['Final'] == 'approved' else
                      'FFEBEE' if row['Final'] == 'rejected' else 'FFFFFF')
                for ci in range(1, len(hdrs)+1):
                    ws.cell(ri, ci).fill = PatternFill('solid', fgColor=bg)
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 16
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf, as_attachment=True,
            download_name=f'applications{label}_{datetime.now().strftime("%Y%m%d")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=20, rightMargin=20, topMargin=30, bottomMargin=20)
        styles = getSampleStyleSheet()
        title_str = 'CIA Retest Applications'
        if year_filter: title_str += f' — Year {year_filter}'
        if section_filter: title_str += f', Section {section_filter}'
        els = [Paragraph(title_str, styles['Title']),
               Paragraph(f'Generated: {datetime.now().strftime("%d %b %Y %H:%M")}',
                         styles['Normal']), Spacer(1, 12)]
        if rows:
            hdrs = ['#', 'Student', 'Reg No', 'Yr', 'Sec', 'Subject', 'Sem', 'CIA',
                    'Type', 'Staff', 'Tutor', 'HOD', 'Final']
            td = [hdrs] + [[str(r['ID']), r['Student'][:16], r['Reg No'],
                             str(r['Year']), str(r['Section']),
                             r['Subject'][:16], str(r['Semester']),
                             str(r['CIA No']), r['Type'],
                             r['Staff'].upper(), r['Tutor'].upper(),
                             r['HOD'].upper(), r['Final'].upper()] for r in rows]
            t = Table(td, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1A237E')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 8),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F5F7FF')]),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ]))
            els.append(t)
        doc.build(els); buf.seek(0)
        return send_file(buf, as_attachment=True,
            download_name=f'applications{label}_{datetime.now().strftime("%Y%m%d")}.pdf',
            mimetype='application/pdf')


# ─── ALL APPLICATIONS (with year/section filter) ──────────────────────────────
@admin_bp.route('/applications')
@login_required
@admin_required
def all_applications():
    year_filter    = request.args.get('year', type=int)
    section_filter = request.args.get('section', '')
    query = RetestApplication.query
    if year_filter:
        query = query.filter_by(student_year=year_filter)
    if section_filter:
        query = query.filter_by(student_section=section_filter)
    apps = query.order_by(RetestApplication.submitted_at.desc()).all()
    year_section_stats = _get_year_section_stats()
    return render_template('admin/all_applications.html', apps=apps,
                           year_section_stats=year_section_stats,
                           sections=SECTIONS,
                           year_filter=year_filter,
                           section_filter=section_filter)


@admin_bp.route('/applications/<int:app_id>')
@login_required
@admin_required
def view_application(app_id):
    app = RetestApplication.query.get_or_404(app_id)
    return render_template('admin/view_application.html', app=app)


# ─── STAFF MANAGEMENT ────────────────────────────────────────────────────────
@admin_bp.route('/staff')
@login_required
@admin_required
def manage_staff():
    staff_list = User.query.filter(
        User.role.in_(['subject_staff','tutor','hod','coordinator'])
    ).order_by(User.name).all()
    return render_template('admin/manage_staff.html', staff_list=staff_list)


@admin_bp.route('/staff/add', methods=['POST'])
@login_required
@admin_required
def add_staff():
    name = request.form.get('name','').strip()
    email = request.form.get('email','').strip().lower()
    phone = request.form.get('phone','').strip()
    role = request.form.get('role','')
    secondary = request.form.get('secondary_role','').strip() or None
    department = request.form.get('department','').strip()
    password = request.form.get('password','staff123').strip() or 'staff123'
    h_year = request.form.get('handling_year','').strip()
    h_sec  = request.form.get('handling_section','').strip().upper()
    if role == 'admin':
        flash('Cannot add Admin here.','danger')
        return redirect(url_for('admin.manage_staff'))
    existing = User.query.filter_by(email=email).first()
    if existing:
        flash('Email exists. Use Edit to change roles.','danger')
        return redirect(url_for('admin.manage_staff'))
    u = User(name=name, email=email, phone=phone, role=role,
             secondary_role=secondary, department=department,
             handling_year=int(h_year) if h_year else None,
             handling_section=h_sec if h_sec in ('A','B','C') else None)
    u.set_password(password)
    db.session.add(u); db.session.commit()
    flash(f'Staff added! Default password: {password}', 'success')
    return redirect(url_for('admin.manage_staff'))


@admin_bp.route('/staff/edit/<int:uid>', methods=['POST'])
@login_required
@admin_required
def edit_staff(uid):
    u = User.query.get_or_404(uid)
    u.name = request.form.get('name', u.name).strip()
    u.phone = request.form.get('phone', u.phone or '').strip()
    u.department = request.form.get('department', u.department or '').strip()
    sec = request.form.get('secondary_role','').strip()
    u.secondary_role = sec if sec else None
    h_year = request.form.get('handling_year','').strip()
    h_sec  = request.form.get('handling_section','').strip().upper()
    u.handling_year = int(h_year) if h_year else None
    u.handling_section = h_sec if h_sec in ('A','B','C') else None
    pw = request.form.get('password','').strip()
    if pw: u.set_password(pw)
    db.session.commit(); flash('Staff updated.','success')
    return redirect(url_for('admin.manage_staff'))


@admin_bp.route('/staff/delete/<int:uid>', methods=['POST'])
@login_required
@admin_required
def delete_staff(uid):
    u = User.query.get_or_404(uid)
    linked = RetestApplication.query.filter(
        (RetestApplication.staff_id == uid) | (RetestApplication.tutor_id == uid)).count()
    if linked:
        flash(f'Cannot delete — linked to {linked} application(s).','danger')
        return redirect(url_for('admin.manage_staff'))
    db.session.delete(u); db.session.commit()
    flash(f'"{u.name}" deleted.', 'success')
    return redirect(url_for('admin.manage_staff'))


@admin_bp.route('/staff/toggle/<int:uid>', methods=['POST'])
@login_required
@admin_required
def toggle_staff(uid):
    u = User.query.get_or_404(uid)
    u.is_active = not u.is_active; db.session.commit()
    flash(f'User {"activated" if u.is_active else "deactivated"}.','success')
    return redirect(url_for('admin.manage_staff'))


# ─── SUBJECT MANAGEMENT ──────────────────────────────────────────────────────
@admin_bp.route('/subjects')
@login_required
@admin_required
def manage_subjects():
    subjects = Subject.query.order_by(Subject.semester, Subject.subject_name).all()
    staff_list = User.query.filter(
        (User.role == 'subject_staff') | (User.secondary_role == 'subject_staff'),
        User.is_active == True).order_by(User.name).all()
    sss_list = SubjectStaffSection.query.order_by(
        SubjectStaffSection.semester, SubjectStaffSection.section).all()
    return render_template('admin/manage_subjects.html',
                           subjects=subjects, staff_list=staff_list,
                           sss_list=sss_list, sections=SECTIONS)


@admin_bp.route('/subjects/add', methods=['POST'])
@login_required
@admin_required
def add_subject():
    sem = int(request.form.get('semester', 1))
    s = Subject(
        subject_name=request.form.get('subject_name','').strip(),
        subject_code=request.form.get('subject_code','').strip(),
        semester=sem,
        year=SEMESTER_TO_YEAR.get(sem, 1),
        department=request.form.get('department','').strip(),
        staff_id=int(request.form.get('staff_id')) if request.form.get('staff_id') else None)
    db.session.add(s); db.session.commit(); flash('Subject added.','success')
    return redirect(url_for('admin.manage_subjects'))


@admin_bp.route('/subjects/edit/<int:sid>', methods=['POST'])
@login_required
@admin_required
def edit_subject(sid):
    s = Subject.query.get_or_404(sid)
    s.subject_name = request.form.get('subject_name', s.subject_name)
    s.subject_code = request.form.get('subject_code', s.subject_code)
    s.semester     = int(request.form.get('semester', s.semester))
    s.year         = SEMESTER_TO_YEAR.get(s.semester, 1)
    s.department   = request.form.get('department', s.department)
    s.staff_id     = int(request.form.get('staff_id')) if request.form.get('staff_id') else None
    s.is_active    = request.form.get('is_active') == 'on'
    db.session.commit(); flash('Subject updated.','success')
    return redirect(url_for('admin.manage_subjects'))


@admin_bp.route('/subjects/delete/<int:sid>', methods=['POST'])
@login_required
@admin_required
def delete_subject(sid):
    s = Subject.query.get_or_404(sid)
    linked = RetestApplication.query.filter_by(subject_id=sid).count()
    if linked:
        flash(f'Cannot delete — {linked} application(s) exist.','danger')
        return redirect(url_for('admin.manage_subjects'))
    SubjectStaffSection.query.filter_by(subject_id=sid).delete()
    CIADate.query.filter_by(subject_id=sid).delete()
    db.session.delete(s); db.session.commit()
    flash('Subject deleted.','success')
    return redirect(url_for('admin.manage_subjects'))


# ─── SECTION-STAFF MAPPING (SubjectStaffSection) ─────────────────────────────
@admin_bp.route('/subjects/section-map/add', methods=['POST'])
@login_required
@admin_required
def add_section_map():
    subject_id = int(request.form.get('subject_id'))
    staff_id   = int(request.form.get('staff_id'))
    semester   = int(request.form.get('semester'))
    section    = request.form.get('section','').upper().strip()
    if section not in ('A','B','C'):
        flash('Invalid section.','danger')
        return redirect(url_for('admin.manage_subjects'))
    existing = SubjectStaffSection.query.filter_by(
        subject_id=subject_id, semester=semester, section=section).first()
    if existing:
        existing.staff_id = staff_id
        flash('Section mapping updated.','success')
    else:
        db.session.add(SubjectStaffSection(
            subject_id=subject_id, staff_id=staff_id,
            semester=semester, section=section))
        flash('Section mapping added.','success')
    db.session.commit()
    return redirect(url_for('admin.manage_subjects'))


@admin_bp.route('/subjects/section-map/delete/<int:mid>', methods=['POST'])
@login_required
@admin_required
def delete_section_map(mid):
    m = SubjectStaffSection.query.get_or_404(mid)
    db.session.delete(m); db.session.commit()
    flash('Section mapping removed.','success')
    return redirect(url_for('admin.manage_subjects'))


# ─── CIA DATES ────────────────────────────────────────────────────────────────
@admin_bp.route('/cia-dates')
@login_required
@admin_required
def manage_cia_dates():
    subjects  = Subject.query.filter_by(is_active=True).order_by(Subject.subject_name).all()
    cia_dates = CIADate.query.order_by(CIADate.exam_date.desc()).all()
    return render_template('admin/manage_cia_dates.html',
                           subjects=subjects, cia_dates=cia_dates, today=date.today())


@admin_bp.route('/cia-dates/add', methods=['POST'])
@login_required
@admin_required
def add_cia_date():
    try:
        ed = datetime.strptime(request.form.get('exam_date'), '%Y-%m-%d').date()
        rs = request.form.get('retest_date','').strip()
        es = request.form.get('application_end_date','').strip()
        rd = datetime.strptime(rs, '%Y-%m-%d').date() if rs else None
        nd = datetime.strptime(es, '%Y-%m-%d').date() if es else None
        db.session.add(CIADate(
            subject_id=int(request.form.get('subject_id')),
            cia_number=int(request.form.get('cia_number')),
            exam_date=ed, retest_date=rd, application_end_date=nd,
            semester=int(request.form.get('semester')),
            academic_year=request.form.get('academic_year','').strip(),
            created_by=current_user.id))
        db.session.commit(); flash('CIA date added.','success')
    except Exception as e:
        flash(f'Error: {e}','danger')
    return redirect(url_for('admin.manage_cia_dates'))


@admin_bp.route('/cia-dates/edit/<int:cid>', methods=['POST'])
@login_required
@admin_required
def edit_cia_date(cid):
    c = CIADate.query.get_or_404(cid)
    try:
        c.exam_date = datetime.strptime(request.form.get('exam_date'), '%Y-%m-%d').date()
        rs = request.form.get('retest_date','').strip()
        es = request.form.get('application_end_date','').strip()
        c.retest_date          = datetime.strptime(rs,'%Y-%m-%d').date() if rs else None
        c.application_end_date = datetime.strptime(es,'%Y-%m-%d').date() if es else None
        c.academic_year        = request.form.get('academic_year', c.academic_year)
        db.session.commit(); flash('CIA date updated.','success')
    except Exception as e:
        flash(f'Error: {e}','danger')
    return redirect(url_for('admin.manage_cia_dates'))


@admin_bp.route('/cia-dates/delete/<int:cid>', methods=['POST'])
@login_required
@admin_required
def delete_cia_date(cid):
    c = CIADate.query.get_or_404(cid)
    db.session.delete(c); db.session.commit()
    flash('CIA date deleted.','success')
    return redirect(url_for('admin.manage_cia_dates'))
from datetime import date, timedelta

@admin_bp.route('/cia-dates/<int:cia_id>/toggle-window', methods=['POST'])
@login_required
@admin_required
def toggle_retest_window(cia_id):
    cia = CIADate.query.get_or_404(cia_id)

    today = date.today()

    # If the window is currently open, close it.
    if cia.is_application_open():
        cia.application_end_date = today - timedelta(days=1)
        flash('Retest application window closed successfully.', 'success')
    else:
        # Open the application window.
        # The exam must be in the past for is_application_open() to return True.
        if not cia.exam_date or cia.exam_date >= today:
            cia.exam_date = today - timedelta(days=1)

        # Keep applications open for the next 7 days.
        cia.application_end_date = today + timedelta(days=7)

        flash('Retest application window opened successfully.', 'success')

    db.session.commit()

    return redirect(url_for('admin.manage_cia_dates'))

# ─── ABSENTEES ───────────────────────────────────────────────────────────────
@admin_bp.route('/absentees')
@login_required
@admin_or_hod_required
def view_absentees():
    year_filter = request.args.get('year', type=int)
    records_cia1 = _absence_records_for_cia(1, year_filter)
    records_cia2 = _absence_records_for_cia(2, year_filter)
    records_cia3 = _absence_records_for_cia(3, year_filter)
    return render_template('admin/absentees.html',
                           records_cia1=records_cia1,
                           records_cia2=records_cia2,
                           records_cia3=records_cia3,
                           year_filter=year_filter)


@admin_bp.route('/absentees/download/cia<int:cia_num>/<fmt>')
@login_required
@admin_or_hod_required
def download_absentees_cia(cia_num, fmt):
    year_filter = request.args.get('year', type=int)
    records = _absence_records_for_cia(cia_num, year_filter)
    rows = []
    for rec in records:
        students = rec.get_students()
        if students:
            for s in students:
                if year_filter and _absence_student_year(rec, s) != year_filter:
                    continue
                rows.append({'Subject': rec.subject.subject_name,
                             'Code': rec.subject.subject_code, 'CIA': cia_num,
                             'Semester': rec.semester or '',
                             'Reg No': s.get('reg_no',''),
                             'Student Name': s.get('name',''),
                             'Uploaded By': rec.uploader.name,
                             'Date': rec.uploaded_at.strftime('%d %b %Y')})
        else:
            rows.append({'Subject': rec.subject.subject_name,
                         'Code': rec.subject.subject_code, 'CIA': cia_num,
                         'Semester': rec.semester or '',
                         'Reg No': '—', 'Student Name': '(file only)',
                         'Uploaded By': rec.uploader.name,
                         'Date': rec.uploaded_at.strftime('%d %b %Y')})
    if fmt == 'excel':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook(); ws = wb.active
        ws.title = f'CIA {cia_num} Absentees'
        hdrs = ['Subject','Code','CIA','Semester','Reg No','Student Name',
                'Uploaded By','Date']
        for ci, h in enumerate(hdrs, 1):
            c = ws.cell(1, ci, h)
            c.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
            c.fill = PatternFill('solid', fgColor='1A237E')
            c.alignment = Alignment(horizontal='center')
        for ri, row in enumerate(rows, 2):
            for ci, v in enumerate([row[h] for h in hdrs], 1):
                ws.cell(ri, ci, v).fill = PatternFill(
                    'solid', fgColor='F5F7FF' if ri % 2 == 0 else 'FFFFFF')
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        label = f'_Year{year_filter}' if year_filter else ''
        return send_file(buf, as_attachment=True,
            download_name=f'absentees_CIA{cia_num}{label}_{datetime.now().strftime("%Y%m%d")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Table,
                                        TableStyle, Paragraph, Spacer)
        from reportlab.lib.styles import getSampleStyleSheet
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
                                leftMargin=30, rightMargin=30,
                                topMargin=40, bottomMargin=30)
        styles = getSampleStyleSheet()
        title = f'Absentee List — CIA {cia_num}'
        if year_filter:
            title += f' | Year {year_filter}'
        els = [Paragraph(title, styles['Title']),
               Paragraph(f'Generated: {datetime.now().strftime("%d %b %Y %H:%M")}',
                         styles['Normal']), Spacer(1, 12)]
        if rows:
            hdrs = ['Subject', 'CIA', 'Sem', 'Reg No', 'Student Name', 'Uploaded By']
            td = [hdrs] + [[r['Subject'][:18], str(r['CIA']),
                             str(r['Semester']), r['Reg No'],
                             r['Student Name'], r['Uploaded By']] for r in rows]
            t = Table(td, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0),(-1,0), colors.HexColor('#1A237E')),
                ('TEXTCOLOR',  (0,0),(-1,0), colors.white),
                ('FONTNAME',   (0,0),(-1,0), 'Helvetica-Bold'),
                ('FONTSIZE',   (0,0),(-1,-1), 9),
                ('GRID',       (0,0),(-1,-1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0,1),(-1,-1),
                 [colors.white, colors.HexColor('#EEF2FF')]),
                ('ALIGN',      (0,0),(-1,-1), 'CENTER'),
            ]))
            els.append(t)
        doc.build(els); buf.seek(0)
        label = f'_Year{year_filter}' if year_filter else ''
        return send_file(buf, as_attachment=True,
            download_name=f'absentees_CIA{cia_num}{label}_{datetime.now().strftime("%Y%m%d")}.pdf',
            mimetype='application/pdf')



# ─── RETRANSMIT ───────────────────────────────────────────────────────────────
@admin_bp.route('/applications/<int:app_id>/retransmit', methods=['POST'])
@login_required
def retransmit(app_id):
    # Allow the approval roles to restart a rejected application from the failed stage.
    allowed = (current_user.role in ('admin', 'hod', 'tutor', 'subject_staff') or
               current_user.secondary_role in ('hod', 'tutor', 'subject_staff'))
    if not allowed:
        flash('Access denied. Only approval staff can retransmit.', 'danger')
        return redirect(url_for('main.index'))

    app = RetestApplication.query.get_or_404(app_id)
    if (current_user.role == 'subject_staff' or current_user.secondary_role == 'subject_staff') and app.staff_id != current_user.id:
        flash('You can retransmit only applications assigned to you.', 'danger')
        return redirect(url_for('staff.dashboard'))

    if app.final_status != 'rejected':
        flash('Only rejected applications can be retransmitted.', 'warning')
        if current_user.role == 'subject_staff' or current_user.secondary_role == 'subject_staff':
            return redirect(url_for('staff.dashboard'))
        if current_user.role == 'tutor' or current_user.secondary_role == 'tutor':
            return redirect(url_for('tutor.dashboard'))
        return redirect(url_for('admin.view_application', app_id=app_id))

    # Reset only the rejected stage
    if app.hod_status == 'rejected':
        app.hod_status = 'pending'; app.hod_remark = None; app.hod_action_time = None
    elif app.coordinator_status == 'rejected':
        app.coordinator_status = 'pending'; app.coordinator_remark = None; app.coordinator_action_time = None
    elif app.tutor_status == 'rejected':
        app.tutor_status = 'pending'; app.tutor_remark = None; app.tutor_action_time = None
    elif app.staff_status == 'rejected':
        app.staff_status = 'pending'; app.staff_remark = None; app.staff_action_time = None

    app.final_status     = 'pending'
    app.retransmit_count = (app.retransmit_count or 0) + 1
    app.retransmit_by    = current_user.id
    app.retransmit_at    = datetime.utcnow()
    db.session.commit()
    flash(f'Application #{app_id} retransmitted successfully.', 'success')

    if current_user.role == 'subject_staff' or current_user.secondary_role == 'subject_staff':
        return redirect(url_for('staff.dashboard'))
    if current_user.role == 'tutor' or current_user.secondary_role == 'tutor':
        return redirect(url_for('tutor.dashboard'))
    return redirect(url_for('admin.view_application', app_id=app_id))


# ─── HELPER ───────────────────────────────────────────────────────────────────
def _expand_register_numbers(raw: str) -> list:
    """
    Expand register-number ranges and comma-separated lists.
    E.g. '6176AC23UCS001-6176AC23UCS028' or
    '2403617610421114 to 2403617610422167' expands to individual reg nos.
    Supports (No Numbers: 19) exclusion annotations.
    """
    if not raw:
        return []
    excluded = set()
    for match in re.finditer(r'\(No Numbers?:\s*(\d+)\)', raw, re.IGNORECASE):
        excluded.add(int(match.group(1)))
    raw = re.sub(r'\(No Numbers?:\s*\d+\)', '', raw, flags=re.IGNORECASE)
    result = []
    parts = re.split(r'[,\n]+', raw)
    for part in parts:
        part = _clean_cell(part)
        if not part:
            continue
        range_match = re.match(r'^(.*?)\s*(?:-|to)\s*(.*?)$', part, re.IGNORECASE)
        if range_match:
            start_reg, end_reg = range_match.groups()
            start_parts = _split_register_tail(start_reg)
            end_parts = _split_register_tail(end_reg)
            if start_parts and end_parts and start_parts[0] == end_parts[0]:
                prefix1, start_str = start_parts
                _, end_str = end_parts
                width = len(start_str)
                start_num = int(start_str)
                end_num = int(end_str)
                if end_num < start_num:
                    start_num, end_num = end_num, start_num
                if (end_num - start_num) > 20000:
                    result.append(_normalize_register_number(part))
                    continue
                for n in range(start_num, end_num + 1):
                    if n not in excluded:
                        result.append(_normalize_register_number(f"{prefix1}{str(n).zfill(width)}"))
            else:
                result.append(_normalize_register_number(part))
        else:
            result.append(_normalize_register_number(part))
    return _dedupe_registers(result)


def _split_register_tail(register_number):
    match = re.match(r'^(.*?)(\d+)$', _clean_cell(register_number))
    if not match:
        return None
    return match.group(1), match.group(2)


def _dedupe_registers(register_numbers):
    seen = set()
    result = []
    for reg in register_numbers:
        reg = _normalize_register_number(reg)
        if reg and reg not in seen:
            seen.add(reg)
            result.append(reg)
    return result


def _registered_student_map(year=None, section=None, active_only=True):
    query = User.query.filter_by(role='student')
    if active_only:
        query = query.filter_by(is_active=True)
    if year:
        query = query.filter_by(year=year)
    if section:
        query = query.filter_by(section=section)
    students = query.all()
    return {
        _normalize_register_number(student.register_number): student
        for student in students
        if _normalize_register_number(student.register_number)
    }


def _resolve_registered_registers(raw_registers, year=None, section=None):
    expanded = _expand_register_numbers(raw_registers)
    if not expanded:
        return [], 0, []
    student_map = _registered_student_map(year=year, section=section, active_only=True)
    resolved = [reg for reg in expanded if reg in student_map]
    return resolved, len(expanded) - len(resolved), expanded


def _find_invigilator(identifier):
    identifier = _clean_cell(identifier)
    if not identifier:
        return None
    query = User.query.filter(
        User.is_active == True,
        (User.role.in_(['subject_staff', 'tutor', 'hod', 'coordinator'])) |
        (User.secondary_role.in_(['subject_staff', 'tutor', 'hod', 'coordinator']))
    )
    if '@' in identifier:
        return query.filter(db.func.lower(User.email) == identifier.lower()).first()
    return query.filter(db.func.lower(User.name) == identifier.lower()).first()


def _hall_student_rows(hall_number, allotments):
    rows = []
    student_map = _registered_student_map(active_only=False)
    attendance = ExamAttendance.query.filter_by(hall_number=hall_number).all()
    attendance_map = {
        (att.seating_id, _normalize_register_number(att.register_number)): att
        for att in attendance
    }

    for allotment in allotments:
        for reg in allotment.get_register_numbers():
            reg = _normalize_register_number(reg)
            student = student_map.get(reg)
            att = attendance_map.get((allotment.id, reg))
            rows.append({
                'seating_id': allotment.id,
                'register_number': reg,
                'student_name': student.name if student else '',
                'year': (student.year if student and student.year else allotment.year),
                'section': (student.section if student and student.section else allotment.section),
                'status': att.status if att else 'present',
                'marked': bool(att),
            })

    if not rows:
        for att in attendance:
            rows.append({
                'seating_id': att.seating_id,
                'register_number': att.register_number,
                'student_name': '',
                'year': att.year,
                'section': att.section,
                'status': att.status,
                'marked': True,
            })
    return rows


# ─── SEATING ALLOTMENT ────────────────────────────────────────────────────────
# --- HALLS AND GENERATED SEATING ------------------------------------------------
YEAR_TEXT = {2: 'II Year', 3: 'III Year', 4: 'IV Year'}
SEAT_POSITIONS = [1,2,3,4,5, 10,9,8,7,6, 11,12,13,14,15, 20,19,18,17,16]
ROW_MAP = {1:1,2:1,3:1,4:1,5:1,6:2,7:2,8:2,9:2,10:2,11:3,12:3,13:3,14:3,15:3,16:4,17:4,18:4,19:4,20:4}
COL_MAP = {1:1,2:2,3:3,4:4,5:5,10:1,9:2,8:3,7:4,6:5,11:1,12:2,13:3,14:4,15:5,20:1,19:2,18:3,17:4,16:5}


def _cia_int(value):
    text = str(value or '').replace('CIA', '').strip()
    return int(text) if text in ('1', '2', '3') else None


def _student_dicts(year):
    return [{'register_number': u.register_number, 'name': u.name, 'department': u.department or ''}
            for u in User.query.filter_by(role='student', year=year, is_active=True)
            .order_by(User.register_number.asc()).all()]


def generate_seating(iv_students, iii_students, hall_id, cia_id, exam_date, generated_by):
    records = []
    for i, pos in enumerate(SEAT_POSITIONS):
        for side, students, label_year, year_name in (
            ('LEFT', iv_students, 'IV YEAR', 'IV Year'),
            ('RIGHT', iii_students, 'III YEAR', 'III Year'),
        ):
            s = students[i] if i < len(students) else None
            records.append({
                'hall_id': hall_id, 'cia_id': cia_id, 'bench_position': pos,
                'seat_side': side, 'seat_label': f'{label_year}-{pos}' if s else f'{label_year}-{pos} (VACANT)',
                'student_reg_no': s['register_number'] if s else None,
                'student_name': s['name'] if s else None,
                'year': year_name, 'department': s['department'] if s else None,
                'row_group': ROW_MAP[pos], 'col_number': COL_MAP[pos],
                'exam_date': exam_date, 'generated_by': generated_by
            })
    return records


def _allocation_groups(cia_id=None, exam_date=None):
    query = SeatingAllocation.query.join(Hall)
    if cia_id:
        query = query.filter(SeatingAllocation.cia_id == cia_id)
    if exam_date:
        query = query.filter(SeatingAllocation.exam_date == exam_date)
    rows = query.order_by(Hall.id, SeatingAllocation.bench_position, SeatingAllocation.seat_side).all()
    groups = {}
    for row in rows:
        groups.setdefault(row.hall_id, {'hall': row.hall, 'rows': []})['rows'].append(row)
    return list(groups.values())


@admin_bp.route('/manage-halls', methods=['GET', 'POST'])
@login_required
@admin_required
def manage_halls():
    if request.method == 'POST':
        action = request.form.get('action', 'add')
        try:
            if action == 'upload':
                f = request.files.get('hall_file')
                if not f or not f.filename.lower().endswith('.xlsx'):
                    flash('Please upload a .xlsx hall file.', 'danger')
                    return redirect(url_for('admin.manage_halls'))
                import pandas as pd
                df = pd.read_excel(f, dtype=str)
                df.columns = [_normalize_column_name(c) for c in df.columns]
                added = updated = duplicates = 0
                for _, row in df.iterrows():
                    hall_number = _row_value(row, 'hall_number')
                    if not hall_number:
                        continue
                    data = {
                        'hall_name': _row_value(row, 'hall_name'),
                        'hall_number': hall_number,
                        'block': _row_value(row, 'block'),
                        'floor': _row_value(row, 'floor'),
                        'capacity': int(float(_row_value(row, 'capacity') or 40)),
                        'is_special': hall_number == '340'
                    }
                    hall = Hall.query.filter_by(hall_number=hall_number).first()
                    if hall:
                        duplicates += 1; updated += 1
                        for key, value in data.items():
                            setattr(hall, key, value)
                    else:
                        db.session.add(Hall(**data)); added += 1
                db.session.commit()
                flash(f'Hall upload complete: {added} added, {updated} updated.', 'success')
                if duplicates:
                    flash(f'{duplicates} duplicate hall number(s) were updated.', 'warning')
            else:
                hall_id = request.form.get('hall_id', type=int)
                hall_number = _clean_cell(request.form.get('hall_number'))
                data = {
                    'hall_name': _clean_cell(request.form.get('hall_name')),
                    'hall_number': hall_number,
                    'block': _clean_cell(request.form.get('block')),
                    'floor': _clean_cell(request.form.get('floor')),
                    'capacity': max(request.form.get('capacity', type=int) or 40, 1),
                    'is_special': bool(request.form.get('is_special')) or hall_number == '340'
                }
                hall = Hall.query.get(hall_id) if hall_id else None
                duplicate = Hall.query.filter(Hall.hall_number == hall_number, Hall.id != (hall.id if hall else 0)).first()
                if duplicate:
                    flash('Duplicate hall number warning: hall already exists.', 'danger')
                elif hall:
                    for key, value in data.items():
                        setattr(hall, key, value)
                    db.session.commit(); flash('Hall updated.', 'success')
                else:
                    db.session.add(Hall(**data)); db.session.commit(); flash('Hall added.', 'success')
        except Exception as e:
            db.session.rollback(); flash(f'Hall save failed: {e}', 'danger')
        return redirect(url_for('admin.manage_halls'))
    return render_template('admin/manage_halls.html', halls=Hall.query.order_by(Hall.id).all())


@admin_bp.route('/manage-halls/<int:hall_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_hall(hall_id):
    db.session.delete(Hall.query.get_or_404(hall_id))
    db.session.commit()
    flash('Hall deleted.', 'success')
    return redirect(url_for('admin.manage_halls'))


@admin_bp.route('/seating-allocation', methods=['GET', 'POST'])
@login_required
@admin_required
def seating_allocation():
    selected_cia = _cia_int(request.values.get('cia_id'))
    selected_date = _parse_date(request.values.get('exam_date', ''))
    if request.method == 'POST':
        cia_id = _cia_int(request.form.get('cia_id'))
        exam_date = _parse_date(request.form.get('exam_date', ''))
        hall_count = request.form.get('hall_count', type=int) or 0
        selected_years = request.form.getlist('years')
        if not cia_id or not exam_date or hall_count < 1:
            flash('CIA number, exam date, and number of halls are required.', 'danger')
            return redirect(url_for('admin.seating_allocation'))
        halls = Hall.query.order_by(Hall.id).limit(hall_count).all()
        auto_halls = [h for h in halls if not h.is_special]
        skipped = [h for h in halls if h.is_special]
        if skipped:
            flash(', '.join([f'Room {h.hall_number} skipped - requires manual seating' for h in skipped]), 'warning')
        if not auto_halls:
            flash('No auto-allocation halls available.', 'danger')
            return redirect(url_for('admin.seating_allocation'))
        if '2' in selected_years:
            flash('II Year selected, but this ACE pattern only seats IV Year left and III Year right.', 'warning')
        iii_students = _student_dicts(3) if '3' in selected_years else []
        iv_students = _student_dicts(4) if '4' in selected_years else []
        SeatingAllocation.query.filter_by(cia_id=cia_id, exam_date=exam_date).delete()
        for idx, hall in enumerate(auto_halls):
            records = generate_seating(iv_students[idx*20:(idx+1)*20], iii_students[idx*20:(idx+1)*20],
                                       hall.id, cia_id, exam_date, current_user.name)
            db.session.add_all([SeatingAllocation(**r) for r in records])
        db.session.commit()
        flash('Seating allocation generated.', 'success')
        return redirect(url_for('admin.seating_allocation', cia_id=cia_id, exam_date=exam_date))
    groups = _allocation_groups(selected_cia, selected_date) if selected_cia and selected_date else []
    return render_template('admin/seating_allocation.html', halls=Hall.query.order_by(Hall.id).all(),
                           groups=groups, selected_cia=selected_cia, selected_date=selected_date)


def _next_vacant(hall_id, cia_id, exam_date, preferred_year):
    side = 'LEFT' if preferred_year == 'IV Year' else 'RIGHT'
    return SeatingAllocation.query.filter_by(hall_id=hall_id, cia_id=cia_id, exam_date=exam_date,
                                             seat_side=side, student_reg_no=None).order_by(SeatingAllocation.bench_position).first()


@admin_bp.route('/update-seat', methods=['POST'])
@login_required
@admin_required
def update_seat():
    row = SeatingAllocation.query.get_or_404(request.form.get('allocation_id', type=int))
    reg_no = _normalize_register_number(request.form.get('student_reg_no', ''))
    hall_id = request.form.get('hall_id', type=int) or row.hall_id
    duplicate = SeatingAllocation.query.filter(SeatingAllocation.id != row.id, SeatingAllocation.hall_id == hall_id,
        SeatingAllocation.cia_id == row.cia_id, SeatingAllocation.exam_date == row.exam_date,
        SeatingAllocation.student_reg_no == reg_no).first() if reg_no else None
    if duplicate:
        return jsonify({'ok': False, 'message': 'Duplicate register number in same hall/CIA/date.'}), 400
    target = _next_vacant(hall_id, row.cia_id, row.exam_date, row.year) if hall_id != row.hall_id else row
    if not target:
        return jsonify({'ok': False, 'message': 'No vacant matching seat in target hall.'}), 400
    if target.id != row.id:
        row.student_reg_no = row.student_name = row.department = None
        row.seat_label = f"{'IV YEAR' if row.year == 'IV Year' else 'III YEAR'}-{row.bench_position} (VACANT)"
    student = User.query.filter_by(role='student', register_number=reg_no).first() if reg_no else None
    target.student_reg_no = reg_no or None
    target.student_name = student.name if student else (_clean_cell(request.form.get('student_name')) or None)
    target.department = student.department if student else target.department
    db.session.commit()
    return jsonify({'ok': True, 'message': 'Seat updated.'})


@admin_bp.route('/swap-hall-seat', methods=['POST'])
@login_required
@admin_required
def swap_hall_seat():
    row = SeatingAllocation.query.get_or_404(request.form.get('allocation_id', type=int))
    target = _next_vacant(request.form.get('hall_id', type=int), row.cia_id, row.exam_date, row.year)
    if not target:
        return jsonify({'ok': False, 'message': 'No vacant matching seat in target hall.'}), 400
    target.student_reg_no, target.student_name, target.department = row.student_reg_no, row.student_name, row.department
    target.seat_label = f"{'IV YEAR' if row.year == 'IV Year' else 'III YEAR'}-{target.bench_position}"
    row.student_reg_no = row.student_name = row.department = None
    row.seat_label = f"{'IV YEAR' if row.year == 'IV Year' else 'III YEAR'}-{row.bench_position} (VACANT)"
    db.session.commit()
    return jsonify({'ok': True, 'message': 'Student moved.'})


@admin_bp.route('/update-exam-date', methods=['POST'])
@login_required
@admin_required
def update_exam_date():
    cia_id = _cia_int(request.form.get('cia_id'))
    old_date = _parse_date(request.form.get('old_date', ''))
    new_date = _parse_date(request.form.get('new_date', ''))
    if not cia_id or not old_date or not new_date:
        return jsonify({'ok': False, 'message': 'CIA, old date, and new date are required.'}), 400
    SeatingAllocation.query.filter_by(cia_id=cia_id, exam_date=old_date).update({'exam_date': new_date})
    db.session.commit()
    return jsonify({'ok': True, 'redirect': url_for('admin.seating_allocation', cia_id=cia_id, exam_date=new_date)})


@admin_bp.route('/regenerate-seating', methods=['POST'])
@login_required
@admin_required
def regenerate_seating():
    cia_id = _cia_int(request.form.get('cia_id'))
    exam_date = _parse_date(request.form.get('exam_date', ''))
    SeatingAllocation.query.filter_by(cia_id=cia_id, exam_date=exam_date).delete()
    db.session.commit()
    flash('Existing allocation cleared. Generate again with the same inputs.', 'success')
    return redirect(url_for('admin.seating_allocation', cia_id=cia_id, exam_date=exam_date))


@admin_bp.route('/regenerate-hall/<int:hall_id>', methods=['POST'])
@login_required
@admin_required
def regenerate_single_hall(hall_id):
    cia_id = _cia_int(request.form.get('cia_id')); exam_date = _parse_date(request.form.get('exam_date', ''))
    hall_ids = [h.id for h in Hall.query.filter_by(is_special=False).order_by(Hall.id).all()]
    offset = hall_ids.index(hall_id) if hall_id in hall_ids else 0
    SeatingAllocation.query.filter_by(hall_id=hall_id, cia_id=cia_id, exam_date=exam_date).delete()
    records = generate_seating(_student_dicts(4)[offset*20:(offset+1)*20], _student_dicts(3)[offset*20:(offset+1)*20],
                               hall_id, cia_id, exam_date, current_user.name)
    db.session.add_all([SeatingAllocation(**r) for r in records]); db.session.commit()
    flash('Single hall regenerated.', 'success')
    return redirect(url_for('admin.seating_allocation', cia_id=cia_id, exam_date=exam_date))


@admin_bp.route('/seating-allocation/download/<fmt>')
@login_required
@admin_required
def download_seating_allocation(fmt):
    cia_id = _cia_int(request.args.get('cia_id')); exam_date = _parse_date(request.args.get('exam_date', ''))
    groups = _allocation_groups(cia_id, exam_date)
    if fmt == 'excel':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        wb = openpyxl.Workbook(); wb.remove(wb.active)
        if not groups:
            ws = wb.create_sheet('No Data')
            ws.append(['No seating allocation found for the selected CIA/date.'])
        for group in groups:
            ws = wb.create_sheet((group['hall'].hall_name or group['hall'].hall_number)[:31])
            ws.merge_cells('A1:F1'); ws['A1'] = 'Adhiyamaan College of Engineering'; ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A2:F2'); ws['A2'] = f'CIA {cia_id} Seating Allocation - {exam_date.strftime("%d-%m-%Y")}'; ws['A2'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A3:F3'); ws['A3'] = f"{group['hall'].hall_name} | Block {group['hall'].block or '-'} | Floor {group['hall'].floor or '-'}"
            ws.append([]); ws.append(['Bench Pos', 'Seat Label', 'Reg No', 'Name', 'Year', 'Department'])
            for c in ws[5]: c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor='1A237E')
            for r in group['rows']:
                ws.append([r.bench_position, r.seat_label, r.student_reg_no or 'VACANT', r.student_name or 'VACANT', r.year or '-', r.department or '-'])
                color = '808080' if not r.student_reg_no else ('1565C0' if r.year == 'IV Year' else 'C62828')
                for c in ws[ws.max_row]: c.font = Font(color=color, strike=not bool(r.student_reg_no))
            for col_idx in range(1, ws.max_column + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 22
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf, as_attachment=True, download_name='seating_allocation.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO(); doc = SimpleDocTemplate(buf, pagesize=landscape(A4)); styles = getSampleStyleSheet(); els = []
    if not groups:
        els.append(Paragraph('No seating allocation found for the selected CIA/date.', styles['Normal']))
    for idx, group in enumerate(groups):
        if idx: els.append(PageBreak())
        els += [Paragraph('Adhiyamaan College of Engineering', styles['Title']),
                Paragraph(f"CIA {cia_id} Seating Allocation - {exam_date.strftime('%d-%m-%Y')}", styles['Heading2']),
                Paragraph(f"Hall: {group['hall'].hall_name} | Block: {group['hall'].block or '-'} | Floor: {group['hall'].floor or '-'}", styles['Normal']), Spacer(1, 10)]
        data = [['Bench Pos', 'Seat Label', 'Reg No', 'Name', 'Year', 'Department']] + [
            [r.bench_position, r.seat_label, r.student_reg_no or 'VACANT', r.student_name or 'VACANT', r.year or '-', r.department or '-']
            for r in group['rows']]
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1A237E')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),0.4,colors.grey),('FONTSIZE',(0,0),(-1,-1),8)]))
        els += [table, Spacer(1, 10), Paragraph(f'Generated by {current_user.name} on {datetime.now().strftime("%d %b %Y %H:%M")}', styles['Normal'])]
    doc.build(els); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='seating_allocation.pdf', mimetype='application/pdf')


@admin_bp.route('/seating')
@login_required
def seating_allotment():
    if not (current_user.role in ('admin', 'hod', 'coordinator') or
            current_user.secondary_role in ('hod', 'coordinator')):
        flash('Access denied.', 'danger')
        return redirect(url_for('main.index'))
    allotments = SeatingAllotment.query.order_by(SeatingAllotment.hall_number).all()
    all_staff  = User.query.filter(
        (User.role.in_(['subject_staff', 'tutor', 'hod', 'coordinator'])) |
        (User.secondary_role.in_(['subject_staff', 'tutor', 'hod', 'coordinator'])),
        User.is_active == True
    ).order_by(User.name).all()
    hall_map = {}
    for a in allotments:
        hall_map.setdefault(a.hall_number, []).append(a)
    return render_template('admin/seating_allotment.html',
                           allotments=allotments, hall_map=hall_map, all_staff=all_staff)


@admin_bp.route('/seating/upload', methods=['POST'])
@login_required
def seating_upload():
    if not (current_user.role in ('admin', 'hod', 'coordinator') or
            current_user.secondary_role in ('hod', 'coordinator')):
        flash('Access denied.', 'danger')
        return redirect(url_for('admin.seating_allotment'))

    added = updated = skipped_rows = skipped_registers = 0
    f = request.files.get('seating_file')
    if f and f.filename:
        ext = f.filename.rsplit('.', 1)[-1].lower()
        if ext not in ('xlsx', 'xls', 'csv'):
            flash('Only Excel/CSV files accepted.', 'danger')
            return redirect(url_for('admin.seating_allotment'))
        try:
            import pandas as pd
            df = pd.read_csv(f) if ext == 'csv' else pd.read_excel(f, dtype=str)
            df.columns = [c.strip().lower().replace(' ', '_') for c in df.columns]
            for _, row in df.iterrows():
                def v(*keys):
                    for k in keys:
                        val = row.get(k, '')
                        if str(val).strip().lower() not in ('nan', 'none', ''):
                            return str(val).strip()
                    return ''
                hall    = v('hall_number', 'hall_no', 'hall')
                exam_dt = _parse_date(v('exam_date', 'seating_date', 'cia_exam_date', 'date'))
                yr_str  = v('year', 'yr')
                sec     = v('section', 'sec').upper()
                reg_raw = v('register_numbers', 'register_number', 'reg_numbers', 'reg_nos')
                num_s   = v('no_of_students', 'num_students', 'count')
                tot_s   = v('total_no_of_students', 'total_students', 'total')
                inv_em  = v('invigilator_email', 'invigilator', 'staff_email', 'invigilator_name')
                if not hall or not exam_dt:
                    skipped_rows += 1
                    continue
                yr_int  = int(float(yr_str)) if yr_str else None
                sec_val = sec if sec in ('A', 'B', 'C') else None
                num_int = int(float(num_s)) if num_s else None
                tot_int = int(float(tot_s)) if tot_s else None
                reg_list, missing_count, _ = _resolve_registered_registers(reg_raw, yr_int, sec_val)
                skipped_registers += missing_count
                if reg_raw and not reg_list:
                    skipped_rows += 1
                    continue
                inv_user = _find_invigilator(inv_em)
                existing = SeatingAllotment.query.filter_by(
                    hall_number=hall, year=yr_int, section=sec_val).first()
                if existing:
                    existing.exam_date = exam_dt or existing.exam_date
                    existing.set_register_numbers(reg_list)
                    existing.num_students   = len(reg_list) or num_int or existing.num_students
                    existing.total_students = len(reg_list) or tot_int or existing.total_students
                    if inv_user:
                        existing.invigilator_id = inv_user.id
                    existing.uploaded_by = current_user.id
                    existing.uploaded_at = datetime.utcnow()
                    updated += 1
                else:
                    sa = SeatingAllotment(
                        hall_number=hall, exam_date=exam_dt, year=yr_int, section=sec_val,
                        num_students=len(reg_list) or num_int,
                        total_students=len(reg_list) or tot_int,
                        invigilator_id=inv_user.id if inv_user else None,
                        uploaded_by=current_user.id)
                    sa.set_register_numbers(reg_list)
                    db.session.add(sa)
                    added += 1
            db.session.commit()
            flash(f'Seating allotment uploaded: {added} new, {updated} updated, '
                  f'{skipped_rows} rows skipped.', 'success')
            if skipped_registers:
                flash(f'{skipped_registers} register number(s) were ignored because they are not active students in the registry/year/section.', 'warning')
        except Exception as e:
            db.session.rollback()
            flash(f'Upload failed: {e}', 'danger')
    elif request.form.get('hall_number'):
        record_id = request.form.get('record_id')
        hall = (request.form.get('hall_number') or '').strip()
        exam_dt = _parse_date(request.form.get('exam_date', ''))
        student_count = request.form.get('student_count') or request.form.get('num_students')
        yr_int = _parse_year(request.form.get('year', ''))
        sec_val = _parse_section(request.form.get('section', ''))
        reg_raw = request.form.get('register_numbers', '')
        invigilator_id = request.form.get('invigilator_id', type=int) or None

        if not hall:
            flash('Hall number is required for manual entry.', 'danger')
            return redirect(url_for('admin.seating_allotment'))
        if not exam_dt:
            flash('Seating / CIA date is required for attendance-to-CIA absentee mapping.', 'danger')
            return redirect(url_for('admin.seating_allotment'))

        reg_list, skipped_registers, _ = _resolve_registered_registers(reg_raw, yr_int, sec_val)
        if reg_raw and not reg_list:
            flash('No active registered students matched that register number list/range.', 'danger')
            return redirect(url_for('admin.seating_allotment'))

        try:
            student_count_int = int(student_count) if student_count else len(reg_list)
        except (TypeError, ValueError):
            student_count_int = len(reg_list)
        student_count_int = len(reg_list) or student_count_int

        if student_count_int < 1:
            flash('At least one registered student or a valid student count is required.', 'danger')
            return redirect(url_for('admin.seating_allotment'))

        if record_id:
            existing = SeatingAllotment.query.get(int(record_id)) if record_id.isdigit() else None
            if not existing:
                flash('Hall record not found.', 'danger')
                return redirect(url_for('admin.seating_allotment'))
            existing.hall_number = hall
            existing.exam_date = exam_dt
            existing.year = yr_int
            existing.section = sec_val
            existing.set_register_numbers(reg_list)
            existing.num_students = student_count_int
            existing.total_students = student_count_int
            existing.invigilator_id = invigilator_id
            existing.uploaded_by = current_user.id
            existing.uploaded_at = datetime.utcnow()
            updated += 1
        else:
            existing = SeatingAllotment.query.filter_by(
                hall_number=hall, year=yr_int, section=sec_val).first()
            if existing:
                existing.exam_date = exam_dt
                existing.year = yr_int
                existing.section = sec_val
                existing.set_register_numbers(reg_list)
                existing.num_students = student_count_int
                existing.total_students = student_count_int
                existing.invigilator_id = invigilator_id
                existing.uploaded_by = current_user.id
                existing.uploaded_at = datetime.utcnow()
                updated += 1
            else:
                sa = SeatingAllotment(
                    hall_number=hall,
                    exam_date=exam_dt,
                    year=yr_int,
                    section=sec_val,
                    num_students=student_count_int,
                    total_students=student_count_int,
                    invigilator_id=invigilator_id,
                    uploaded_by=current_user.id
                )
                sa.set_register_numbers(reg_list)
                db.session.add(sa)
                added += 1
        db.session.commit()
        flash(f'Hall entry saved: {added} new, {updated} updated.', 'success')
        if skipped_registers:
            flash(f'{skipped_registers} register number(s) were ignored because they are not active students in the selected year/section.', 'warning')
    else:
        flash('Please upload a seating file or submit manual hall details.', 'danger')
    return redirect(url_for('admin.seating_allotment'))


@admin_bp.route('/seating/<int:sa_id>/get')
@login_required
def get_seating_entry(sa_id):
    if not (current_user.role in ('admin', 'hod', 'coordinator') or
            current_user.secondary_role in ('hod', 'coordinator')):
        flash('Access denied.', 'danger')
        return redirect(url_for('admin.seating_allotment'))
    sa = SeatingAllotment.query.get_or_404(sa_id)
    return jsonify({
        'id': sa.id,
        'hall_number': sa.hall_number,
        'year': sa.year,
        'section': sa.section,
        'register_numbers': sa.get_register_numbers(),
        'num_students': sa.num_students,
        'total_students': sa.total_students,
        'invigilator_id': sa.invigilator_id
    })


@admin_bp.route('/seating/<int:sa_id>/assign', methods=['POST'])
@login_required
def assign_invigilator(sa_id):
    if not (current_user.role in ('admin', 'hod') or current_user.secondary_role == 'hod'):
        flash('Access denied.', 'danger')
        return redirect(url_for('admin.seating_allotment'))
    sa = SeatingAllotment.query.get_or_404(sa_id)
    sa.invigilator_id = request.form.get('invigilator_id', type=int) or None
    db.session.commit()
    flash('Invigilator assigned.', 'success')
    return redirect(url_for('admin.seating_allotment'))


@admin_bp.route('/seating/<int:sa_id>/delete', methods=['POST'])
@login_required
@admin_required
def delete_seating(sa_id):
    sa = SeatingAllotment.query.get_or_404(sa_id)
    db.session.delete(sa)
    db.session.commit()
    flash('Seating row deleted.', 'success')
    return redirect(url_for('admin.seating_allotment'))


# ─── ATTENDANCE ENTRY (invigilator / staff) ───────────────────────────────────
@admin_bp.route('/attendance/mark/<string:hall_number>')
@login_required
def mark_attendance_page(hall_number):
    allotments = SeatingAllotment.query.filter_by(hall_number=hall_number).order_by(
        SeatingAllotment.year, SeatingAllotment.section).all()
    if not allotments:
        flash('No seating data found for this hall.', 'warning')
        return redirect(url_for('main.index'))
    can_access = (
        current_user.role in ('admin', 'hod', 'coordinator', 'tutor', 'subject_staff') or
        current_user.secondary_role in ('hod', 'coordinator', 'tutor', 'subject_staff')
    )
    if not can_access:
        flash('You do not have access to mark halls.', 'danger')
        return redirect(url_for('main.index'))
    assigned_invigilators = {a.invigilator_id for a in allotments if a.invigilator_id}
    is_exam_admin = (
        current_user.role in ('admin', 'hod', 'coordinator') or
        current_user.secondary_role in ('hod', 'coordinator')
    )
    if assigned_invigilators and not is_exam_admin and current_user.id not in assigned_invigilators:
        flash('This hall is assigned to another invigilator.', 'danger')
        return redirect(url_for('admin.my_halls'))
    seating = allotments[0]
    student_rows = _hall_student_rows(hall_number, allotments)
    planned_count = sum(_allotment_student_count(a) for a in allotments)
    return render_template('admin/mark_attendance.html',
                           hall_number=hall_number,
                           allotments=allotments,
                           seating=seating,
                           student_rows=student_rows,
                           planned_count=planned_count)


@admin_bp.route('/attendance/mark/<string:hall_number>/submit', methods=['POST'])
@login_required
def submit_attendance(hall_number):
    allotments = SeatingAllotment.query.filter_by(hall_number=hall_number).all()
    can_access = (
        current_user.role in ('admin', 'hod', 'coordinator', 'tutor', 'subject_staff') or
        current_user.secondary_role in ('hod', 'coordinator', 'tutor', 'subject_staff')
    )
    if not can_access:
        flash('Access denied.', 'danger')
        return redirect(url_for('main.index'))
    if not allotments:
        flash('No seating data found for this hall.', 'warning')
        return redirect(url_for('main.index'))

    assigned_invigilators = {a.invigilator_id for a in allotments if a.invigilator_id}
    is_exam_admin = (
        current_user.role in ('admin', 'hod', 'coordinator') or
        current_user.secondary_role in ('hod', 'coordinator')
    )
    if assigned_invigilators and not is_exam_admin and current_user.id not in assigned_invigilators:
        flash('This hall is assigned to another invigilator.', 'danger')
        return redirect(url_for('admin.my_halls'))

    allotment_map = {a.id: a for a in allotments}
    saved = absent_count = present_count = 0
    submitted_keys = set()
    absent_regs = {
        _normalize_register_number(reg)
        for reg in request.form.getlist('absent_register[]')
    }
    reg_numbers = request.form.getlist('register_number[]')
    seating_ids = request.form.getlist('seating_id[]')
    years = request.form.getlist('year[]')
    sections = request.form.getlist('section[]')

    for idx, reg in enumerate(reg_numbers):
        reg = _normalize_register_number(reg)
        if not reg:
            continue
        try:
            seating_id = int(seating_ids[idx]) if idx < len(seating_ids) else 0
        except ValueError:
            seating_id = 0
        seating = allotment_map.get(seating_id)
        if not seating:
            continue
        year_raw = years[idx] if idx < len(years) else ''
        section = (sections[idx] if idx < len(sections) else '').strip().upper()
        try:
            year = int(year_raw) if year_raw else None
        except ValueError:
            year = None
        if section not in ('A', 'B', 'C', 'D', 'E'):
            section = None
        status = 'absent' if reg in absent_regs else 'present'
        submitted_keys.add((seating.id, reg))
        att = ExamAttendance.query.filter_by(seating_id=seating.id, register_number=reg).first()
        if att:
            att.exam_date = seating.exam_date
            att.year = year
            att.section = section
            att.status = status
            att.marked_by = current_user.id
            att.marked_at = datetime.utcnow()
        else:
            att = ExamAttendance(
                seating_id=seating.id, hall_number=seating.hall_number,
                exam_date=seating.exam_date,
                register_number=reg, year=year, section=section,
                status=status, marked_by=current_user.id)
            db.session.add(att)
        saved += 1
        if status == 'absent':
            absent_count += 1
        else:
            present_count += 1

    existing_records = ExamAttendance.query.filter(
        ExamAttendance.seating_id.in_(list(allotment_map.keys()))
    ).all()
    for record in existing_records:
        key = (record.seating_id, _normalize_register_number(record.register_number))
        if key not in submitted_keys:
            db.session.delete(record)

    db.session.commit()
    flash(f'Attendance saved for hall {hall_number}: {absent_count} absent, {present_count} present ({saved} students).', 'success')
    is_adm_hod = (current_user.role in ('admin', 'hod') or current_user.secondary_role == 'hod')
    if is_adm_hod:
        return redirect(url_for('admin.view_attendance'))
    return redirect(url_for('admin.mark_attendance_page', hall_number=hall_number))


# ─── ATTENDANCE OVERVIEW (admin/HOD) ─────────────────────────────────────────
# --- GENERATED ATTENDANCE SUMMARY ---------------------------------------------
@admin_bp.route('/attendance-summary')
@login_required
@admin_required
def attendance_summary():
    cia_id = _cia_int(request.args.get('cia_id')) or 1
    exam_date = _parse_date(request.args.get('exam_date', '')) or date.today()
    year_filter = request.args.get('year', '')
    dept_filter = request.args.get('department', '')
    groups = _attendance_summary_data(cia_id, exam_date, year_filter, dept_filter)
    departments = [d[0] for d in db.session.query(SeatingAllocation.department).filter(SeatingAllocation.department != None).distinct().all()]
    return render_template('admin/attendance_summary.html', groups=groups, cia_id=cia_id,
                           exam_date=exam_date, year_filter=year_filter, dept_filter=dept_filter,
                           departments=sorted([d for d in departments if d]))


def _attendance_summary_data(cia_id, exam_date, year_filter='', dept_filter=''):
    if not cia_id or not exam_date:
        return []
    halls = db.session.query(Hall).join(SeatingAllocation).filter(
        SeatingAllocation.cia_id == cia_id,
        SeatingAllocation.exam_date == exam_date,
        SeatingAllocation.student_reg_no != None
    ).group_by(Hall.id).order_by(Hall.id).all()
    groups = []
    for hall in halls:
        q = SeatingAllocation.query.filter_by(cia_id=cia_id, exam_date=exam_date, hall_id=hall.id).filter(SeatingAllocation.student_reg_no != None)
        if year_filter:
            q = q.filter_by(year=year_filter)
        if dept_filter:
            q = q.filter_by(department=dept_filter)
        students = q.order_by(SeatingAllocation.bench_position, SeatingAllocation.seat_side).all()
        detail, present, absent, invigilator = [], 0, 0, ''
        for s in students:
            att = HallAttendance.query.filter_by(cia_id=cia_id, hall_id=hall.id, exam_date=exam_date, student_reg_no=s.student_reg_no).first()
            status = att.status if att else 'Pending'
            if att and att.invigilator:
                invigilator = att.invigilator.name
            present += 1 if status == 'Present' else 0
            absent += 1 if status == 'Absent' else 0
            detail.append({'seat': s, 'status': status})
        marked = HallAttendance.query.filter_by(cia_id=cia_id, hall_id=hall.id, exam_date=exam_date).count() > 0
        total = len(students)
        groups.append({'hall': hall, 'total': total, 'present': present, 'absent': absent,
                       'pct': round((present / total * 100), 1) if total else 0,
                       'status': 'Marked' if marked else 'Pending',
                       'invigilator': invigilator, 'detail': detail})
    return groups


@admin_bp.route('/attendance-summary/download')
@login_required
@admin_required
def download_generated_attendance_summary():
    cia_id = _cia_int(request.args.get('cia_id')) or 1
    exam_date = _parse_date(request.args.get('exam_date', '')) or date.today()
    groups = _attendance_summary_data(cia_id, exam_date, request.args.get('year', ''), request.args.get('department', ''))
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = 'Hall Summary'
    ws.append(['Hall Name', 'Hall No', 'Total', 'Present', 'Absent', '%', 'Status'])
    detail = wb.create_sheet('Student Detail')
    detail.append(['Hall Name', 'Bench Pos', 'Seat Label', 'Reg No', 'Name', 'Year', 'Dept', 'Status'])
    for cell in ws[1] + detail[1]:
        cell.font = Font(bold=True, color='FFFFFF'); cell.fill = PatternFill('solid', fgColor='1A237E')
    for g in groups:
        ws.append([g['hall'].hall_name, g['hall'].hall_number, g['total'], g['present'], g['absent'], g['pct'], g['status']])
        if g['status'] == 'Pending':
            for c in ws[ws.max_row]: c.fill = PatternFill('solid', fgColor='FFF59D')
        for d in g['detail']:
            s = d['seat']; detail.append([g['hall'].hall_name, s.bench_position, s.seat_label, s.student_reg_no, s.student_name, s.year, s.department, d['status']])
            fill = 'C8E6C9' if d['status'] == 'Present' else ('FFCDD2' if d['status'] == 'Absent' else 'FFF59D')
            for c in detail[detail.max_row]: c.fill = PatternFill('solid', fgColor=fill)
    for sheet in (ws, detail):
        for col in sheet.columns: sheet.column_dimensions[col[0].column_letter].width = 20
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='attendance_summary.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@admin_bp.route('/attendance')
@login_required
def view_attendance():
    if not (current_user.role in ('admin', 'hod') or current_user.secondary_role == 'hod'):
        flash('Access denied.', 'danger')
        return redirect(url_for('main.index'))
    hall_filter    = request.args.get('hall', '')
    section_filter = request.args.get('section', '')
    year_filter    = request.args.get('year', type=int)
    records, hall_summary, ys_summary, all_halls = _attendance_report_data(
        hall_filter, section_filter, year_filter)
    return render_template('admin/attendance.html',
                           records=records, hall_summary=hall_summary,
                           ys_summary=ys_summary, all_halls=all_halls,
                           hall_filter=hall_filter, section_filter=section_filter,
                           year_filter=year_filter, sections=SECTIONS)


def _attendance_report_data(hall_filter='', section_filter='', year_filter=None):
    query = ExamAttendance.query
    if hall_filter:    query = query.filter_by(hall_number=hall_filter)
    if section_filter: query = query.filter_by(section=section_filter)
    if year_filter:    query = query.filter_by(year=year_filter)
    records   = query.order_by(ExamAttendance.hall_number, ExamAttendance.register_number).all()
    attendance_halls = {h[0] for h in db.session.query(ExamAttendance.hall_number).distinct().all()}
    seating_halls = {h[0] for h in db.session.query(SeatingAllotment.hall_number).distinct().all()}
    all_halls = sorted(attendance_halls | seating_halls)
    summary_halls = [hall_filter] if hall_filter else all_halls
    hall_summary = {}
    for h in summary_halls:
        hrs = ExamAttendance.query.filter_by(hall_number=h)
        seating_q = SeatingAllotment.query.filter_by(hall_number=h)
        if section_filter:
            hrs = hrs.filter_by(section=section_filter)
            seating_q = seating_q.filter_by(section=section_filter)
        if year_filter:
            hrs = hrs.filter_by(year=year_filter)
            seating_q = seating_q.filter_by(year=year_filter)
        planned = sum(
            _allotment_student_count(a)
            for a in seating_q.all()
        )
        marked_total = hrs.count()
        hall_summary[h] = {
            'total':   marked_total or planned,
            'present': hrs.filter_by(status='present').count(),
            'absent':  hrs.filter_by(status='absent').count(),
            'marked':  marked_total,
        }
    ys_summary = {}
    for yr in range(1, 5):
        ys_summary[yr] = {}
        for sec in SECTIONS:
            recs = ExamAttendance.query.filter_by(year=yr, section=sec)
            if hall_filter:
                recs = recs.filter_by(hall_number=hall_filter)
            if section_filter and sec != section_filter:
                recs = recs.filter(ExamAttendance.id == None)
            if year_filter and yr != year_filter:
                recs = recs.filter(ExamAttendance.id == None)
            ys_summary[yr][sec] = {
                'total':   recs.count(),
                'present': recs.filter_by(status='present').count(),
                'absent':  recs.filter_by(status='absent').count(),
            }
    return records, hall_summary, ys_summary, all_halls


@admin_bp.route('/attendance/download/<fmt>')
@login_required
def download_attendance_summary(fmt):
    if not (current_user.role in ('admin', 'hod') or current_user.secondary_role == 'hod'):
        flash('Access denied.', 'danger')
        return redirect(url_for('main.index'))
    if fmt not in ('excel', 'pdf'):
        flash('Unsupported download format.', 'danger')
        return redirect(url_for('admin.view_attendance'))

    hall_filter    = request.args.get('hall', '')
    section_filter = request.args.get('section', '')
    year_filter    = request.args.get('year', type=int)
    _, hall_summary, ys_summary, _ = _attendance_report_data(
        hall_filter, section_filter, year_filter)

    hall_rows = []
    for hall, stats in hall_summary.items():
        pct = round((stats['present'] / stats['total'] * 100), 1) if stats['total'] else 0
        hall_rows.append({
            'Hall No.': hall,
            'Total Students': stats['total'],
            'Present': stats['present'],
            'Absent': stats['absent'],
            'Attendance %': pct,
        })

    year_rows = []
    for yr in range(1, 5):
        row = {'Year': ['','I','II','III','IV'][yr]}
        year_total = year_absent = 0
        for sec in SECTIONS:
            d = ys_summary[yr][sec]
            row[f'Section {sec}'] = (
                f"Present {d['present']} / Absent {d['absent']} / Total {d['total']}"
                if d['total'] else '-'
            )
            year_total += d['total']
            year_absent += d['absent']
        row['Year Total'] = f'{year_total} total, {year_absent} absent' if year_total else '-'
        year_rows.append(row)

    label = ''
    if hall_filter: label += f'_Hall{hall_filter}'
    if year_filter: label += f'_Year{year_filter}'
    if section_filter: label += f'_Section{section_filter}'
    today_label = datetime.now().strftime('%Y%m%d')

    if fmt == 'excel':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Hall Summary'
        _write_summary_sheet(ws, hall_rows)
        ys = wb.create_sheet('Year Section Summary')
        _write_summary_sheet(ys, year_rows)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True,
            download_name=f'attendance_summary{label}_{today_label}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=24, rightMargin=24, topMargin=30, bottomMargin=24)
    styles = getSampleStyleSheet()
    filters = []
    if hall_filter: filters.append(f'Hall {hall_filter}')
    if year_filter: filters.append(f'Year {year_filter}')
    if section_filter: filters.append(f'Section {section_filter}')
    title = 'Attendance Summary'
    if filters:
        title += ' | ' + ', '.join(filters)
    els = [Paragraph(title, styles['Title']),
           Paragraph(f'Generated: {datetime.now().strftime("%d %b %Y %H:%M")}', styles['Normal']),
           Spacer(1, 12)]
    els.append(Paragraph('Hall-wise Summary', styles['Heading2']))
    els.append(_pdf_table(
        [['Hall No.', 'Total Students', 'Present', 'Absent', 'Attendance %']] +
        [[r['Hall No.'], r['Total Students'], r['Present'], r['Absent'], f"{r['Attendance %']}%"]
         for r in hall_rows],
        colors.HexColor('#1A237E')
    ))
    els.append(Spacer(1, 14))
    els.append(Paragraph('Year / Section Summary', styles['Heading2']))
    ys_headers = ['Year'] + [f'Section {s}' for s in SECTIONS] + ['Year Total']
    els.append(_pdf_table(
        [ys_headers] + [[r[h] for h in ys_headers] for r in year_rows],
        colors.HexColor('#0277BD')
    ))
    doc.build(els)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
        download_name=f'attendance_summary{label}_{today_label}.pdf',
        mimetype='application/pdf')


def _write_summary_sheet(ws, rows):
    from openpyxl.styles import Font, PatternFill, Alignment
    if not rows:
        ws.cell(1, 1, 'No data')
        return
    hdrs = list(rows[0].keys())
    for ci, h in enumerate(hdrs, 1):
        cell = ws.cell(1, ci, h)
        cell.font = Font(bold=True, color='FFFFFF', name='Arial')
        cell.fill = PatternFill('solid', fgColor='1A237E')
        cell.alignment = Alignment(horizontal='center')
    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(hdrs, 1):
            ws.cell(ri, ci, row[h])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 24


def _pdf_table(data, header_color):
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), header_color),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F5F7FF')]),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    return table


# ─── MY HALLS (invigilator) ──────────────────────────────────────────────────
@admin_bp.route('/my-halls')
@login_required
def my_halls():
    can_view_all = (
        current_user.role in ('admin', 'hod', 'coordinator') or
        current_user.secondary_role in ('hod', 'coordinator')
    )
    if can_view_all:
        halls = SeatingAllotment.query.order_by(SeatingAllotment.hall_number).all()
    else:
        halls = SeatingAllotment.query.filter(
            (SeatingAllotment.invigilator_id == None) |
            (SeatingAllotment.invigilator_id == current_user.id)
        ).order_by(SeatingAllotment.hall_number).all()
    hall_numbers = sorted({h.hall_number for h in halls})
    hall_status = {}
    for hn in hall_numbers:
        rows = SeatingAllotment.query.filter_by(hall_number=hn).all()
        if not can_view_all:
            rows = [r for r in rows if r.invigilator_id in (None, current_user.id)]
        total = sum(_allotment_student_count(r) for r in rows)
        seating_ids = [r.id for r in rows]
        marked = ExamAttendance.query.filter(
            ExamAttendance.seating_id.in_(seating_ids)
        ).count() if seating_ids else 0
        hall_status[hn] = {'total': total, 'marked': marked}
    return render_template('admin/my_halls.html',
                           halls=halls, hall_numbers=hall_numbers, hall_status=hall_status)


# ─── SUBJECT-WISE ABSENTEES (admin view) ─────────────────────────────────────
@admin_bp.route('/absentees/subject-wise')
@login_required
def view_absentees_subject():
    if not (current_user.role in ('admin', 'hod') or current_user.secondary_role == 'hod'):
        flash('Access denied.', 'danger')
        return redirect(url_for('main.index'))
    cia_filter     = request.args.get('cia', type=int)
    section_filter = request.args.get('section', '')
    query = AbsenceRecord.query
    if cia_filter:
        query = query.filter_by(cia_number=cia_filter)
    records = query.order_by(AbsenceRecord.cia_number, AbsenceRecord.subject_id).all()
    tree = {}
    for rec in records:
        uploader = rec.uploader
        section  = uploader.handling_section if uploader else '?'
        if section_filter and section != section_filter:
            continue
        key = (rec.subject.subject_name, rec.subject.subject_code)
        tree.setdefault(key, {}).setdefault(rec.cia_number, []).append({
            'rec': rec, 'section': section, 'count': len(rec.get_students())
        })
    return render_template('admin/absentees_subject.html',
                           tree=tree, cia_filter=cia_filter,
                           section_filter=section_filter, sections=SECTIONS)
