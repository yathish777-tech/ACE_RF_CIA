import io
from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user
from models import db, RetestApplication, CIADate
from datetime import datetime, date
from functools import wraps

coordinator_bp = Blueprint('coordinator', __name__)

def coordinator_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not (current_user.role == 'coordinator' or
                current_user.secondary_role == 'coordinator'):
            flash('Access denied.', 'danger')
            return redirect(url_for('main.index'))
        return f(*args, **kwargs)
    return decorated


@coordinator_bp.route('/dashboard')
@login_required
@coordinator_required
def dashboard():
    year_filter    = request.args.get('year', type=int)
    section_filter = request.args.get('section', '')

    pending_q = RetestApplication.query.filter_by(
        submission_type='late', staff_status='approved',
        tutor_status='approved', coordinator_status='pending', final_status='pending')
    reviewed_q = RetestApplication.query.filter(
        RetestApplication.coordinator_status.in_(['approved','rejected']))

    if year_filter:
        pending_q  = pending_q.filter_by(student_year=year_filter)
        reviewed_q = reviewed_q.filter_by(student_year=year_filter)
    if section_filter:
        pending_q  = pending_q.filter_by(student_section=section_filter)
        reviewed_q = reviewed_q.filter_by(student_section=section_filter)

    pending  = pending_q.order_by(RetestApplication.submitted_at.desc()).all()
    reviewed = reviewed_q.order_by(RetestApplication.coordinator_action_time.desc()).all()

    upcoming_dates = CIADate.query.filter(
        CIADate.application_end_date >= date.today()
    ).order_by(CIADate.application_end_date).limit(5).all()

    stats = {
        'pending' : len(pending),
        'approved': sum(1 for r in reviewed if r.coordinator_status == 'approved'),
        'rejected': sum(1 for r in reviewed if r.coordinator_status == 'rejected'),
        'total'   : len(pending) + len(reviewed),
    }
    return render_template('coordinator/dashboard.html',
                           pending=pending, reviewed=reviewed,
                           stats=stats, upcoming_dates=upcoming_dates,
                           year_filter=year_filter, section_filter=section_filter,
                           sections=['A', 'B', 'C'])


@coordinator_bp.route('/action/<int:app_id>', methods=['POST'])
@login_required
@coordinator_required
def action(app_id):
    application = RetestApplication.query.get_or_404(app_id)
    act    = request.form.get('action')
    remark = request.form.get('remark', '').strip()
    application.coordinator_status      = 'approved' if act == 'approve' else 'rejected'
    application.coordinator_remark      = remark
    application.coordinator_action_time = datetime.utcnow()
    if act == 'reject':
        application.final_status = 'rejected'
        db.session.commit()
        try:
            from utils.email_utils import notify_student_final
            notify_student_final(application)
        except: pass
        flash('Rejected. Student notified.', 'warning')
    else:
        db.session.commit()
        try:
            from utils.email_utils import notify_hod_after_coordinator_late
            notify_hod_after_coordinator_late(application)
        except: pass
        flash('Approved. Forwarded to HOD.', 'success')
    return redirect(url_for('coordinator.dashboard'))


# ─── DOWNLOAD ─────────────────────────────────────────────────────────────────
@coordinator_bp.route('/applications/download/<fmt>')
@login_required
@coordinator_required
def download_applications(fmt):
    year_filter    = request.args.get('year', type=int)
    section_filter = request.args.get('section', '')

    query = RetestApplication.query.filter(
        RetestApplication.coordinator_status.in_(['approved','rejected','pending']))
    if year_filter:    query = query.filter_by(student_year=year_filter)
    if section_filter: query = query.filter_by(student_section=section_filter)
    apps = query.order_by(RetestApplication.submitted_at.desc()).all()

    rows = [{'ID': a.id, 'Student': a.student_name, 'Reg No': a.register_number,
             'Year': a.student_year or '', 'Section': a.student_section or '',
             'Subject': a.subject.subject_name, 'Semester': a.semester,
             'CIA No': a.cia_number, 'Type': a.submission_type.upper(),
             'Staff': a.staff_status, 'Tutor': a.tutor_status,
             'Coordinator': a.coordinator_status, 'HOD': a.hod_status,
             'Final': a.final_status,
             'Submitted': a.submitted_at.strftime('%d %b %Y %H:%M')} for a in apps]

    label = ''
    if year_filter:    label += f'_Year{year_filter}'
    if section_filter: label += f'_Sec{section_filter}'

    if fmt == 'excel':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Applications'
        if rows:
            hdrs = list(rows[0].keys())
            for ci, h in enumerate(hdrs, 1):
                c = ws.cell(1, ci, h)
                c.font = Font(bold=True, color='FFFFFF', name='Arial')
                c.fill = PatternFill('solid', fgColor='1A237E')
                c.alignment = Alignment(horizontal='center')
            for ri, row in enumerate(rows, 2):
                for ci, v in enumerate(row.values(), 1): ws.cell(ri, ci, v)
                bg = ('E8F5E9' if row['Final']=='approved' else
                      'FFEBEE' if row['Final']=='rejected' else 'FFFFFF')
                for ci in range(1, len(hdrs)+1):
                    ws.cell(ri, ci).fill = PatternFill('solid', fgColor=bg)
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 15
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf, as_attachment=True,
            download_name=f'coordinator_applications{label}_{datetime.now().strftime("%Y%m%d")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=20, rightMargin=20, topMargin=30, bottomMargin=20)
        styles = getSampleStyleSheet()
        title_str = 'Coordinator — CIA Retest Applications'
        if year_filter:    title_str += f' | Year {year_filter}'
        if section_filter: title_str += f' | Section {section_filter}'
        els = [Paragraph(title_str, styles['Title']),
               Paragraph(f'Generated: {datetime.now().strftime("%d %b %Y %H:%M")}', styles['Normal']),
               Spacer(1,12)]
        if rows:
            hdrs = ['#','Student','Reg No','Yr','Sec','Subject','Sem','CIA','Type','Coord','HOD','Final']
            td = [hdrs]+[[str(r['ID']),r['Student'][:14],r['Reg No'],str(r['Year']),
                           str(r['Section']),r['Subject'][:14],str(r['Semester']),
                           str(r['CIA No']),r['Type'],r['Coordinator'].upper(),
                           r['HOD'].upper(),r['Final'].upper()] for r in rows]
            t = Table(td, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1A237E')),
                ('TEXTCOLOR',(0,0),(-1,0),colors.white),
                ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
                ('FONTSIZE',(0,0),(-1,-1),8),
                ('GRID',(0,0),(-1,-1),0.5,colors.grey),
                ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F5F7FF')]),
                ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ]))
            els.append(t)
        doc.build(els); buf.seek(0)
        return send_file(buf, as_attachment=True,
            download_name=f'coordinator_applications{label}_{datetime.now().strftime("%Y%m%d")}.pdf',
            mimetype='application/pdf')
