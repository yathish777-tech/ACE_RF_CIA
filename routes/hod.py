import io
from flask import Blueprint, render_template, redirect, url_for, flash, request, send_file
from flask_login import login_required, current_user
from models import db, RetestApplication, CIADate, User
from datetime import datetime, date
from functools import wraps

hod_bp = Blueprint('hod', __name__)
SECTIONS = ['A', 'B', 'C']

def hod_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        # Allow both HOD and Admin access
        if not (current_user.role in ('hod', 'admin') or current_user.secondary_role == 'hod'):
            flash('Access denied.', 'danger')
            return redirect(url_for('main.index'))
        return f(*args, **kwargs)
    return decorated


def _get_year_section_stats():
    apps = RetestApplication.query.all()
    stats = {}
    for year in range(1, 5):
        stats[year] = {}
        for sec in SECTIONS:
            year_apps = [a for a in apps if (a.student_year == year and a.student_section == sec)]
            stats[year][sec] = {
                'total': len(year_apps),
                'approved': sum(1 for a in year_apps if a.final_status == 'approved'),
                'rejected': sum(1 for a in year_apps if a.final_status == 'rejected'),
                'pending': sum(1 for a in year_apps if a.final_status == 'pending'),
                'apps': year_apps
            }
    return stats


@hod_bp.route('/dashboard')
@login_required
@hod_required
def dashboard():
    year_filter    = request.args.get('year', type=int)
    section_filter = request.args.get('section', '')

    pending_pre  = RetestApplication.query.filter_by(
        submission_type='pre', staff_status='approved',
        tutor_status='approved', hod_status='pending', final_status='pending').all()
    pending_late = RetestApplication.query.filter_by(
        submission_type='late', staff_status='approved', tutor_status='approved',
        coordinator_status='approved', hod_status='pending', final_status='pending').all()
    pending = pending_pre + pending_late

    # Apply year/section filter to reviewed
    reviewed_q = RetestApplication.query.filter(
        RetestApplication.hod_status.in_(['approved','rejected']))
    if year_filter:
        reviewed_q = reviewed_q.filter_by(student_year=year_filter)
    if section_filter:
        reviewed_q = reviewed_q.filter_by(student_section=section_filter)
    reviewed = reviewed_q.order_by(RetestApplication.hod_action_time.desc()).all()

    upcoming_dates = CIADate.query.filter(
        CIADate.application_end_date >= date.today()
    ).order_by(CIADate.application_end_date).limit(5).all()
    all_cia = CIADate.query.order_by(CIADate.exam_date.desc()).all()

    year_section_stats = _get_year_section_stats()

    stats = {
        'pending': len(pending),
        'approved': sum(1 for r in reviewed if r.hod_status == 'approved'),
        'rejected': sum(1 for r in reviewed if r.hod_status == 'rejected'),
        'total': len(pending) + len(reviewed)
    }
    return render_template('hod/dashboard.html',
                           pending=pending, reviewed=reviewed, stats=stats,
                           upcoming_dates=upcoming_dates, all_cia=all_cia,
                           today=date.today(), year_section_stats=year_section_stats,
                           sections=SECTIONS, year_filter=year_filter,
                           section_filter=section_filter)


@hod_bp.route('/action/<int:app_id>', methods=['POST'])
@login_required
@hod_required
def action(app_id):
    application = RetestApplication.query.get_or_404(app_id)
    act    = request.form.get('action')
    remark = request.form.get('remark','').strip()
    application.hod_status = 'approved' if act == 'approve' else 'rejected'
    application.hod_remark = remark
    application.hod_action_time = datetime.utcnow()
    if act == 'reject':
        application.final_status = 'rejected'
        # For pre-submission, HOD is the final reviewer.
        # Do not mark coordinator_status here or it may cause retransmit to restart from the wrong stage.
        db.session.commit()
        try:
            from utils.email_utils import notify_student_final
            notify_student_final(application)
        except: pass
        flash('Rejected. Student notified.', 'warning')
    else:
        application.final_status = 'approved'
        db.session.commit()
        try:
            from utils.email_utils import notify_student_final
            notify_student_final(application)
        except: pass
        flash('FINAL APPROVAL done. Student notified.', 'success')
    return redirect(url_for('hod.dashboard'))


@hod_bp.route('/retransmit/<int:app_id>', methods=['POST'])
@login_required
@hod_required
def retransmit(app_id):
    app = RetestApplication.query.get_or_404(app_id)
    if app.final_status != 'rejected':
        flash('Only rejected applications can be retransmitted.', 'warning')
        return redirect(url_for('hod.dashboard'))
    if app.hod_status == 'rejected':
        app.hod_status = 'pending'; app.hod_remark = None; app.hod_action_time = None
    elif app.coordinator_status == 'rejected':
        app.coordinator_status = 'pending'; app.coordinator_remark = None; app.coordinator_action_time = None
    elif app.tutor_status == 'rejected':
        app.tutor_status = 'pending'; app.tutor_remark = None; app.tutor_action_time = None
    elif app.staff_status == 'rejected':
        app.staff_status = 'pending'; app.staff_remark = None; app.staff_action_time = None
    app.final_status = 'pending'
    if hasattr(app, 'retransmit_count'):
        app.retransmit_count = (getattr(app, 'retransmit_count', 0) or 0) + 1
    db.session.commit()
    flash(f'Application #{app_id} retransmitted.', 'success')
    return redirect(url_for('hod.dashboard'))


@hod_bp.route('/toggle-window/<int:cid>', methods=['POST'])
@login_required
@hod_required
def toggle_window(cid):
    cia = CIADate.query.get_or_404(cid)
    cia.application_window_open = not cia.application_window_open
    db.session.commit()
    flash(f'Retest window {"opened" if cia.application_window_open else "closed"} for CIA {cia.cia_number} — {cia.subject.subject_name}.', 'success')
    return redirect(url_for('hod.dashboard'))


# ─── DOWNLOAD APPLICATIONS ───────────────────────────────────────────────────
@hod_bp.route('/applications/download/<fmt>')
@login_required
@hod_required
def download_applications(fmt):
    year_filter    = request.args.get('year', type=int)
    section_filter = request.args.get('section', '')
    query = RetestApplication.query
    if year_filter:
        query = query.filter_by(student_year=year_filter)
    if section_filter:
        query = query.filter_by(student_section=section_filter)
    apps = query.order_by(RetestApplication.submitted_at.desc()).all()

    rows = [{'ID': a.id, 'Student': a.student_name, 'Reg No': a.register_number,
             'Year': a.student_year or '', 'Section': a.student_section or '',
             'Subject': a.subject.subject_name, 'Semester': a.semester,
             'CIA No': a.cia_number, 'Type': a.submission_type.upper(),
             'Staff': a.staff_status, 'Tutor': a.tutor_status,
             'HOD': a.hod_status, 'Final': a.final_status,
             'Submitted': a.submitted_at.strftime('%d %b %Y %H:%M')} for a in apps]

    label = ''
    if year_filter: label += f'_Year{year_filter}'
    if section_filter: label += f'_Sec{section_filter}'

    if fmt == 'excel':
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Applications'
        if rows:
            hdrs = list(rows[0].keys())
            for ci, h in enumerate(hdrs, 1):
                c = ws.cell(1, ci, h)
                c.font = Font(bold=True, color='FFFFFF', name='Arial')
                c.fill = PatternFill('solid', fgColor='1A237E')
                c.alignment = Alignment(horizontal='center')
            for ri, row in enumerate(rows, 2):
                for ci, v in enumerate(row.values(), 1):
                    ws.cell(ri, ci, v)
                bg = ('E8F5E9' if row['Final'] == 'approved' else
                      'FFEBEE' if row['Final'] == 'rejected' else 'FFFFFF')
                for ci in range(1, len(hdrs)+1):
                    ws.cell(ri, ci).fill = PatternFill('solid', fgColor=bg)
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 16
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return send_file(buf, as_attachment=True,
            download_name=f'hod_applications{label}_{datetime.now().strftime("%Y%m%d")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                leftMargin=20, rightMargin=20, topMargin=30, bottomMargin=20)
        styles = getSampleStyleSheet()
        title_str = 'HOD — CIA Retest Applications'
        if year_filter: title_str += f' | Year {year_filter}'
        if section_filter: title_str += f' | Section {section_filter}'
        els = [Paragraph(title_str, styles['Title']),
               Paragraph(f'Generated: {datetime.now().strftime("%d %b %Y %H:%M")}',
                         styles['Normal']), Spacer(1, 12)]
        if rows:
            hdrs = ['#', 'Student', 'Reg No', 'Yr', 'Sec', 'Subject', 'Sem', 'CIA', 'Type', 'Staff', 'Tutor', 'HOD', 'Final']
            td = [hdrs] + [[str(r['ID']), r['Student'][:14], r['Reg No'],
                             str(r['Year']), str(r['Section']),
                             r['Subject'][:14], str(r['Semester']),
                             str(r['CIA No']), r['Type'],
                             r['Staff'].upper(), r['Tutor'].upper(),
                             r['HOD'].upper(), r['Final'].upper()] for r in rows]
            t = Table(td, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1A237E')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 8),
                ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F5F7FF')]),
                ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ]))
            els.append(t)
        doc.build(els); buf.seek(0)
        return send_file(buf, as_attachment=True,
            download_name=f'hod_applications{label}_{datetime.now().strftime("%Y%m%d")}.pdf',
            mimetype='application/pdf')
